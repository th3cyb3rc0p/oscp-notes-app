"""OSCP Notes - native macOS note-taking app for OSCP exam prep."""
from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import sys
import traceback
import webbrowser
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import markdown as md_lib
from pygments import highlight
from pygments.formatters import HtmlFormatter
from pygments.lexers import get_lexer_by_name, TextLexer
from pygments.util import ClassNotFound

APP_NAME = "OSCP Notes"
APP_VERSION = "1.1"
APP_AUTHOR = "Mayur Parmar (th3cyb3rc0p)"
APP_LINKEDIN = "https://www.linkedin.com/in/th3cyb3rc0p/"

# ----------------------------- platform detection ----------------------------- #
# Used to pick fonts, keybindings, accelerators, and the data-folder location.
IS_MAC = sys.platform == "darwin"
IS_WIN = sys.platform.startswith("win")
IS_LINUX = sys.platform.startswith("linux")


def _user_data_dir() -> Path:
	r"""Resolve the per-user data folder in a platform-appropriate way.

	macOS: ~/OSCP-Notes (preserves v1.0/v1.1 location)
	Linux: $XDG_DATA_HOME/OSCP-Notes, else ~/.local/share/OSCP-Notes
	Windows: %APPDATA%/OSCP-Notes  (e.g. C:\Users\<u>\AppData\Roaming\OSCP-Notes)
	"""
	if IS_MAC:
		return Path.home() / "OSCP-Notes"
	if IS_WIN:
		base = os.environ.get("APPDATA") or str(Path.home() / "AppData" / "Roaming")
		return Path(base) / "OSCP-Notes"
	# Linux / other Unix
	base = os.environ.get("XDG_DATA_HOME") or str(Path.home() / ".local" / "share")
	return Path(base) / "OSCP-Notes"


APP_DIR = _user_data_dir()
DATA_DIR = APP_DIR / "data"
DB_PATH = DATA_DIR / "notes.db"

# Platform-appropriate font families. Tk falls back gracefully if a name is missing.
MONO_FONT = "Menlo" if IS_MAC else ("Consolas" if IS_WIN else "DejaVu Sans Mono")
UI_FONT = "Helvetica Neue" if IS_MAC else ("Segoe UI" if IS_WIN else "DejaVu Sans")


def _accel(label: str) -> str:
	"""Rewrite 'Cmd' to 'Ctrl' in accelerator labels for non-macOS."""
	return label if IS_MAC else label.replace("Cmd", "Ctrl")

PHASES = ["enum", "exploit", "privesc", "post", "cheatsheet", "ad", "general"]
PHASE_LABELS = {
	"enum": "Enumeration",
	"exploit": "Exploitation",
	"privesc": "Privilege Escalation",
	"post": "Post-Exploitation",
	"cheatsheet": "Cheat Sheets",
	"ad": "Active Directory",
	"general": "General",
}

CODE_LANG_MAP = {
	"py": "python", "python": "python", "python3": "python",
	"sh": "bash", "bash": "bash", "shell": "bash", "zsh": "bash",
	"ps1": "powershell", "powershell": "powershell", "ps": "powershell",
	"cmd": "batch", "bat": "batch",
	"php": "php",
	"js": "javascript", "ts": "typescript",
	"html": "html", "css": "css",
	"json": "json", "xml": "xml", "sql": "sql",
	"c": "c", "cpp": "cpp", "csharp": "csharp",
	"java": "java", "go": "go", "ruby": "ruby", "perl": "perl",
	"yaml": "yaml", "yml": "yaml", "ini": "ini",
	"asm": "x86asm", "nasm": "x86asm",
}

PYGMENTS_CACHE = {}

# Theme registry. Each theme is a complete palette. Add a new key here and
# the theme becomes available in View -> Theme. The first theme is the default.
THEMES = {
	"OSCP Dark": {
		"BG": "#1e1e1e", "BG_ALT": "#252526",
		"FG": "#d4d4d4", "FG_DIM": "#858585",
		"ACCENT": "#0e7490", "ACCENT_HOVER": "#0891b2",
		"SELECT_BG": "#264f78", "SELECT_FG": "#ffffff",
		"CODE_BG": "#0d0d0d", "CODE_FG": "#ce9178",
		"BORDER": "#3e3e42",
		"H1": "#569cd6", "H2": "#569cd6", "H3": "#9cdcfe", "H4": "#dcdcaa",
		"LINK": "#569cd6", "QUOTE": "#9cdcfe",
		"PYGMENTS_STYLE": "monokai",
	},
	"Hack The Box": {
		"BG": "#0a0e14", "BG_ALT": "#11151c",
		"FG": "#a3b3c2", "FG_DIM": "#5c6773",
		"ACCENT": "#9fef00", "ACCENT_HOVER": "#b6ff1a",
		"SELECT_BG": "#1a2a0a", "SELECT_FG": "#9fef00",
		"CODE_BG": "#03060a", "CODE_FG": "#9fef00",
		"BORDER": "#1a2230",
		"H1": "#9fef00", "H2": "#9fef00", "H3": "#b6ff1a", "H4": "#ff7b00",
		"LINK": "#9fef00", "QUOTE": "#ff7b00",
		"PYGMENTS_STYLE": "monokai",
	},
	"TryHackMe": {
		"BG": "#1c2536", "BG_ALT": "#222e44",
		"FG": "#e6e6e6", "FG_DIM": "#8a9bb4",
		"ACCENT": "#f3c300", "ACCENT_HOVER": "#ffd133",
		"SELECT_BG": "#2c3a5a", "SELECT_FG": "#ffffff",
		"CODE_BG": "#0e1626", "CODE_FG": "#ff8a65",
		"BORDER": "#2c3a5a",
		"H1": "#f3c300", "H2": "#f3c300", "H3": "#88ccff", "H4": "#88ccff",
		"LINK": "#88ccff", "QUOTE": "#88ccff",
		"PYGMENTS_STYLE": "monokai",
	},
	"Light": {
		"BG": "#ffffff", "BG_ALT": "#f3f3f3",
		"FG": "#1f1f1f", "FG_DIM": "#6a6a6a",
		"ACCENT": "#0969da", "ACCENT_HOVER": "#218bff",
		"SELECT_BG": "#cce5ff", "SELECT_FG": "#1f1f1f",
		"CODE_BG": "#f6f8fa", "CODE_FG": "#953800",
		"BORDER": "#d0d7de",
		"H1": "#0550ae", "H2": "#0550ae", "H3": "#6639ba", "H4": "#953800",
		"LINK": "#0969da", "QUOTE": "#1a7f37",
		"PYGMENTS_STYLE": "friendly",
	},
}

DEFAULT_THEME = "OSCP Dark"

# Module-level convenience references (updated by _apply_theme).
BG = THEMES[DEFAULT_THEME]["BG"]
BG_ALT = THEMES[DEFAULT_THEME]["BG_ALT"]
FG = THEMES[DEFAULT_THEME]["FG"]
FG_DIM = THEMES[DEFAULT_THEME]["FG_DIM"]
ACCENT = THEMES[DEFAULT_THEME]["ACCENT"]
ACCENT_HOVER = THEMES[DEFAULT_THEME]["ACCENT_HOVER"]
SELECT_BG = THEMES[DEFAULT_THEME]["SELECT_BG"]
SELECT_FG = THEMES[DEFAULT_THEME]["SELECT_FG"]
CODE_BG = THEMES[DEFAULT_THEME]["CODE_BG"]
CODE_FG = THEMES[DEFAULT_THEME]["CODE_FG"]
BORDER = THEMES[DEFAULT_THEME]["BORDER"]


def pygments_formatter():
	style = THEMES.get(_ACTIVE_THEME, THEMES[DEFAULT_THEME])["PYGMENTS_STYLE"]
	key = f"formatter_{style}"
	if key not in PYGMENTS_CACHE:
		PYGMENTS_CACHE[key] = HtmlFormatter(
			style=style,
			noclasses=True,
			nobackground=True,
		)
	return PYGMENTS_CACHE[key]


_ACTIVE_THEME = DEFAULT_THEME


def render_markdown(text):
	def replace_code(m):
		lang = m.group(1).strip().lower() if m.group(1) else ""
		lang = CODE_LANG_MAP.get(lang, lang)
		code = m.group(2)
		try:
			lexer = get_lexer_by_name(lang) if lang else TextLexer()
		except ClassNotFound:
			lexer = TextLexer()
		return highlight(code, lexer, pygments_formatter())

	text2 = re.sub(r"```(\w*)\n(.*?)```", replace_code, text, flags=re.DOTALL)
	html = md_lib.markdown(
		text2,
		extensions=["fenced_code", "tables", "sane_lists", "nl2br"],
		output_format="html5",
	)
	return html


# ----------------------------- export helpers --------------------------------- #

def export_as_markdown(row) -> bytes:
	body = f"# {row['title']}\n\n<!-- phase: {row['phase']} - tags: {row['tags']} -->\n\n{row['content']}\n"
	return body.encode("utf-8")


def export_as_text(row) -> bytes:
	"""Strip markdown - keep code blocks verbatim, drop other formatting."""
	out = []
	in_code = False
	for line in row["content"].splitlines():
		if line.strip().startswith("```"):
			in_code = not in_code
			out.append("─" * 60 if not in_code else "")
			continue
		if in_code:
			out.append(line)
			continue
		# strip headings, list markers, emphasis
		clean = re.sub(r"^#{1,6}\s+", "", line)
		clean = re.sub(r"^\s*[-*+]\s+", " - ", clean)
		clean = re.sub(r"\*\*([^*]+)\*\*", r"\1", clean)
		clean = re.sub(r"(?<!\*)\*([^*\n]+)\*(?!\*)", r"\1", clean)
		clean = re.sub(r"`([^`]+)`", r"\1", clean)
		out.append(clean)
	return ("\n".join(out) + "\n").encode("utf-8")


def export_as_html(row, theme_css="") -> bytes:
	html_body = render_markdown(row["content"])
	full = (
		"<!DOCTYPE html>\n<html><head><meta charset='utf-8'><title>"
		+ row["title"].replace("<", "&lt;") + "</title>\n"
		+ f"<style>{theme_css}</style>"
		+ "</head><body>\n"
		+ f"<h1>{row['title']}</h1>\n"
		+ f"<p><em>phase: {row['phase']} - tags: {row['tags']}</em></p>\n"
		+ html_body + "\n</body></html>"
	)
	return full.encode("utf-8")


def export_as_json(row) -> bytes:
	data = {
		"id": row["id"], "title": row["title"], "content": row["content"],
		"phase": row["phase"], "tags": row["tags"],
		"created_at": row["created_at"], "updated_at": row["updated_at"],
	}
	return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")


def export_as_csv(notes: list) -> bytes:
	"""One row per note, includes a `content` column for full text."""
	buf = io.StringIO()
	w = csv.writer(buf)
	w.writerow(["id", "title", "phase", "tags", "created_at", "updated_at", "content"])
	for n in notes:
		w.writerow([n["id"], n["title"], n["phase"], n["tags"], n["created_at"], n["updated_at"], n["content"]])
	return buf.getvalue().encode("utf-8")


def export_as_xlsx(notes: list, path: str):
	"""Write a single-sheet .xlsx with all notes."""
	import openpyxl
	from openpyxl.styles import Font, PatternFill, Alignment
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Notes"
	headers = ["ID", "Title", "Phase", "Tags", "Created", "Updated", "Content"]
	ws.append(headers)
	for cell in ws[1]:
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill("solid", fgColor="1F2937")
	ws.column_dimensions["A"].width = 6
	ws.column_dimensions["B"].width = 40
	ws.column_dimensions["C"].width = 14
	ws.column_dimensions["D"].width = 20
	ws.column_dimensions["E"].width = 20
	ws.column_dimensions["F"].width = 20
	ws.column_dimensions["G"].width = 100
	for n in notes:
		ws.append([n["id"], n["title"], n["phase"], n["tags"], n["created_at"], n["updated_at"], n["content"]])
	wb.save(path)


def export_as_pdf(notes: list, path: str, title="OSCP Notes Export"):
	"""Render a list of notes as a paginated PDF using ReportLab."""
	from reportlab.lib.pagesizes import LETTER
	from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
	from reportlab.lib.units import inch
	from reportlab.lib import colors
	from reportlab.platypus import (
		SimpleDocTemplate, Paragraph, Spacer, Preformatted, PageBreak,
	)
	from reportlab.lib.enums import TA_LEFT
	from reportlab.pdfbase import pdfmetrics
	from reportlab.pdfbase.ttfonts import TTFont
	from reportlab.pdfbase.pdfmetrics import registerFont

	# Try to register a unicode TTF for code blocks; fall back to Courier.
	# Pick the first available font for the current OS so the PDF looks right
	# on macOS (Menlo), Windows (Consolas), and Linux (DejaVu Sans Mono).
	try:
		candidates = [
			("/System/Library/Fonts/Menlo.ttc", "Menlo", 0),
			(r"C:\Windows\Fonts\consola.ttf", "Consolas", None),
			(r"C:\Windows\Fonts\cour.ttf", "Courier New", None),
			("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", "DejaVuSansMono", None),
			("/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf", "DejaVuSansMono-Bold", None),
			("/usr/share/fonts/TTF/DejaVuSansMono.ttf", "DejaVuSansMono", None),
			("/usr/share/fonts/dejavu/DejaVuSansMono.ttf", "DejaVuSansMono", None),
		]
		code_font = "Courier"
		for path_str, font_name, subfont in candidates:
			if Path(path_str).exists():
				try:
					if subfont is not None:
						registerFont(TTFont(font_name, path_str, subfontIndex=subfont))
					else:
						registerFont(TTFont(font_name, path_str))
					code_font = font_name
					break
				except Exception:
					continue
	except Exception:
		code_font = "Courier"

	doc = SimpleDocTemplate(
		path, pagesize=LETTER,
		leftMargin=0.7*inch, rightMargin=0.7*inch,
		topMargin=0.7*inch, bottomMargin=0.7*inch,
		title=title,
	)
	styles = getSampleStyleSheet()
	h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, spaceAfter=8, textColor=colors.HexColor("#0e7490"))
	h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14, spaceAfter=6, textColor=colors.HexColor("#0e7490"))
	body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=13, spaceAfter=4)
	meta = ParagraphStyle("meta", parent=body, textColor=colors.HexColor("#6b7280"), fontSize=8, spaceAfter=10)
	code = ParagraphStyle("code", parent=body, fontName=code_font, fontSize=8, leading=10, backColor=colors.HexColor("#0d0d0d"), textColor=colors.HexColor("#ce9178"), leftIndent=8, rightIndent=8, spaceBefore=4, spaceAfter=4, borderPadding=4)

	flow = []
	flow.append(Paragraph(title, h1))
	flow.append(Spacer(1, 0.15*inch))
	flow.append(Paragraph(f"Generated: {datetime.now().isoformat(timespec='seconds')}", meta))

	for n in notes:
		flow.append(Paragraph(n["title"].replace("<","&lt;").replace(">","&gt;"), h1))
		flow.append(Paragraph(f"phase: {n['phase']} &nbsp; tags: {n['tags']}", meta))
		# Parse the markdown crudely: fenced code blocks -> Preformatted; lines outside -> Paragraph
		in_code = False
		buf: list[str] = []
		for line in n["content"].splitlines():
			if line.strip().startswith("```"):
				if in_code:
					# close
					flow.append(Preformatted("\n".join(buf), code))
					buf = []
					in_code = False
				else:
					if buf:
						# flush text block
						flow.append(Paragraph(_md_to_html("\n".join(buf)), body))
						buf = []
					in_code = True
				continue
			if in_code:
				buf.append(line)
			else:
				buf.append(line)
		if buf:
			if in_code:
				flow.append(Preformatted("\n".join(buf), code))
			else:
				flow.append(Paragraph(_md_to_html("\n".join(buf)), body))
		flow.append(PageBreak())

	doc.build(flow)


def _md_to_html(text: str) -> str:
	"""Tiny markdown -> reportlab-friendly HTML for PDF export."""
	# escape
	s = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
	# headings
	s = re.sub(r"^######\s+(.+)$", r"<font size='9'><b>\1</b></font>", s, flags=re.M)
	s = re.sub(r"^#####\s+(.+)$", r"<font size='10'><b>\1</b></font>", s, flags=re.M)
	s = re.sub(r"^####\s+(.+)$", r"<b>\1</b><br/>", s, flags=re.M)
	s = re.sub(r"^###\s+(.+)$", r"<b>\1</b><br/>", s, flags=re.M)
	s = re.sub(r"^##\s+(.+)$", r"<b><font size='12'>\1</font></b><br/>", s, flags=re.M)
	s = re.sub(r"^#\s+(.+)$", r"<b><font size='14'>\1</font></b><br/>", s, flags=re.M)
	# bold
	s = re.sub(r"\*\*([^*]+)\*\*", r"<b>\1</b>", s)
	# italic
	s = re.sub(r"(?<!\*)\*([^*\n]+)\*(?!\*)", r"<i>\1</i>", s)
	# inline code
	s = re.sub(r"`([^`]+)`", r"<font name='Courier'>\1</font>", s)
	# lists
	s = re.sub(r"^(\s*)[-*+]\s+", r"\1• ", s, flags=re.M)
	# paragraphs (blank line -> break)
	s = s.replace("\n\n", "<br/><br/>")
	s = s.replace("\n", "<br/>")
	return s


# ----------------------------- encryption --------------------------------- #

def derive_vault_key(password: str, salt: bytes) -> bytes:
	"""Derive a Fernet-compatible key from a password using scrypt."""
	dk = hashlib.scrypt(
		password.encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32,
	)
	return base64.urlsafe_b64encode(dk)


def vault_encrypt(plaintext: str, password: str, salt: bytes) -> bytes:
	from cryptography.fernet import Fernet
	key = derive_vault_key(password, salt)
	return Fernet(key).encrypt(plaintext.encode("utf-8"))


def vault_decrypt(ciphertext: bytes, password: str, salt: bytes) -> str:
	from cryptography.fernet import Fernet, InvalidToken
	key = derive_vault_key(password, salt)
	return Fernet(key).decrypt(ciphertext).decode("utf-8")


# ----------------------------- report generator --------------------------------- #

FLAG_PATTERNS = [
	re.compile(r"\bflag\{[^}]+\}", re.IGNORECASE),
	re.compile(r"\bHTB\{[^}]+\}", re.IGNORECASE),
	re.compile(r"\bTHM\{[^}]+\}", re.IGNORECASE),
	re.compile(r"\bOS-\d+\{[^}]+\}", re.IGNORECASE),
	re.compile(r"\b[0-9a-f]{32}\b"),
]
USER_RE = re.compile(r"(?im)^\s*(?:user(?:name)?|login|u)\s*[:=]\s*(\S+)")
PASS_RE = re.compile(r"(?im)^\s*(?:pass(?:word)?|p|secret)\s*[:=]\s*(\S+)")
HASH_RE = re.compile(r"\b[0-9a-fA-F]{32}(?::[0-9a-fA-F]{32})?\b")


def generate_report_shell(row) -> str:
	title = row["title"]
	content = row["content"] or ""

	# Headings
	headings = re.findall(r"^#{1,6}\s+(.+)$", content, re.M)
	# Code blocks
	code_blocks = re.findall(r"```[a-zA-Z0-9_]*\n(.*?)```", content, re.DOTALL)
	# Inline code samples (commands)
	inline_cmds = re.findall(r"`([^`\n]+)`", content)
	# Flags
	flags = set()
	for pat in FLAG_PATTERNS:
		flags.update(pat.findall(content))
	# User / pass / hash lines
	users = USER_RE.findall(content)
	passwords = PASS_RE.findall(content)
	hashes = HASH_RE.findall(content)

	now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

	out = []
	out.append(f"# Penetration Test Report - {title}")
	out.append("")
	out.append(f"_Generated: {now}_")
	out.append("")
	out.append("## 1. Executive Summary")
	out.append("<!-- 2-3 sentence summary of the machine, attack chain, and outcome. -->")
	out.append("")
	out.append("## 2. Service Enumeration")
	if code_blocks:
		out.append("Commands run during enumeration:")
		out.append("```bash")
		for cb in code_blocks[:6]:
			out.append(cb.strip())
			out.append("")
		out.append("```")
	else:
		out.append("<!-- paste nmap / gobuster / ffuf / feroxbuster output here -->")
		out.append("")
	out.append("## 3. Initial Access")
	out.append("- Vulnerability:")
	out.append("- Exploit:")
	if code_blocks:
		out.append("- Command(s):")
		out.append("```bash")
		out.append(code_blocks[0].strip() if code_blocks else "# exploit command here")
		out.append("```")
	out.append("")
	out.append("## 4. Privilege Escalation")
	out.append("- Vector:")
	out.append("- Command(s):")
	out.append("```bash")
	out.append("# privesc command here")
	out.append("```")
	out.append("")
	out.append("## 5. Loot / Credentials")
	if users or passwords or hashes:
		out.append("| user | password | hash | source |")
		out.append("|------|----------|------|--------|")
		for u, p in zip(users, passwords + [""] * (len(users) - len(passwords))):
			out.append(f"| {u} | {p} | | |")
		for h in hashes[:5]:
			out.append(f"| | | `{h}` | |")
	else:
		out.append("<!-- table of captured credentials and hashes -->")
	out.append("")
	out.append("## 6. Flags")
	if flags:
		for f in sorted(flags):
			out.append(f"- `{f}`")
	else:
		out.append("- user.txt: ")
		out.append("- root.txt: ")
	out.append("")
	out.append("## 7. Methodology Highlights")
	if headings:
		out.append("Notes structure:")
		for h in headings[:10]:
			out.append(f"- {h}")
	out.append("")
	out.append("## 8. Timeline")
	out.append("| time | action |")
	out.append("|------|--------|")
	out.append("| 00:00 | Started enumeration |")
	out.append("")
	out.append("## 9. Screenshots")
	out.append("<!-- paste screenshot references or use the Tracker tab to maintain a list -->")
	out.append("")
	out.append("## 10. Lessons Learned")
	out.append("<!-- what you would do differently, what missed, what to remember -->")
	out.append("")
	return "\n".join(out)


class Store:
	def __init__(self, db_path):
		self.db_path = db_path
		self.db_path.parent.mkdir(parents=True, exist_ok=True)
		self.conn = sqlite3.connect(str(db_path))
		self.conn.row_factory = sqlite3.Row
		self._init_schema()

	def _init_schema(self):
		c = self.conn.cursor()
		c.execute("""
			CREATE TABLE IF NOT EXISTS prefs (
				key TEXT PRIMARY KEY,
				value TEXT NOT NULL
			)
		""")
		c.execute("""
			CREATE TABLE IF NOT EXISTS notes (
				id INTEGER PRIMARY KEY AUTOINCREMENT,
				title TEXT NOT NULL,
				content TEXT NOT NULL,
				phase TEXT NOT NULL,
				tags TEXT NOT NULL DEFAULT '',
				created_at TEXT NOT NULL,
				updated_at TEXT NOT NULL
			)
		""")
		c.execute("""
			CREATE VIRTUAL TABLE IF NOT EXISTS notes_fts USING fts5(
				title, content, tags, content='notes', content_rowid='id',
				tokenize='porter unicode61'
			)
		""")
		c.execute("""
			CREATE TRIGGER IF NOT EXISTS notes_ai AFTER INSERT ON notes BEGIN
				INSERT INTO notes_fts(rowid,title,content,tags)
				VALUES (new.id,new.title,new.content,new.tags);
			END;
		""")
		c.execute("""
			CREATE TRIGGER IF NOT EXISTS notes_ad AFTER DELETE ON notes BEGIN
				DELETE FROM notes_fts WHERE rowid=old.id;
			END;
		""")
		c.execute("""
			CREATE TRIGGER IF NOT EXISTS notes_au AFTER UPDATE ON notes BEGIN
				UPDATE notes_fts SET title=new.title, content=new.content, tags=new.tags
				WHERE rowid=new.id;
			END;
		""")
		# Exam-day tracker
		c.execute("""CREATE TABLE IF NOT EXISTS machines (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			name TEXT NOT NULL UNIQUE,
			address TEXT DEFAULT '',
			difficulty TEXT DEFAULT 'unknown',
			started_at TEXT,
			paused_secs INTEGER DEFAULT 0,
			status TEXT DEFAULT 'active',
			notes TEXT DEFAULT '',
			flags TEXT DEFAULT '',
			created_at TEXT, updated_at TEXT
		)""")
		c.execute("""CREATE TABLE IF NOT EXISTS machine_creds (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			machine_id INTEGER REFERENCES machines(id) ON DELETE CASCADE,
			username TEXT, password TEXT, secret TEXT,
			source TEXT, found_at TEXT
		)""")
		c.execute("""CREATE TABLE IF NOT EXISTS machine_loot (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			machine_id INTEGER REFERENCES machines(id) ON DELETE CASCADE,
			name TEXT, kind TEXT, value TEXT,
			captured_at TEXT
		)""")
		c.execute("""CREATE TABLE IF NOT EXISTS machine_screens (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			machine_id INTEGER REFERENCES machines(id) ON DELETE CASCADE,
			label TEXT, path TEXT, taken_at TEXT
		)""")
		c.execute("""CREATE TABLE IF NOT EXISTS machine_checklist (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			machine_id INTEGER REFERENCES machines(id) ON DELETE CASCADE,
			phase TEXT, item TEXT, ticked INTEGER DEFAULT 0,
			UNIQUE(machine_id, phase, item)
		)""")
		# Payloads
		c.execute("""CREATE TABLE IF NOT EXISTS payloads (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			name TEXT NOT NULL,
			category TEXT, platform TEXT,
			content TEXT, tags TEXT DEFAULT '',
			created_at TEXT
		)""")
		c.execute("""CREATE VIRTUAL TABLE IF NOT EXISTS payloads_fts USING fts5(
			name, content, tags, content='payloads', content_rowid='id',
			tokenize='porter unicode61'
		)""")
		c.execute("""CREATE TRIGGER IF NOT EXISTS payloads_ai AFTER INSERT ON payloads BEGIN
			INSERT INTO payloads_fts(rowid,name,content,tags)
			VALUES (new.id,new.name,new.content,new.tags);
		END;""")
		c.execute("""CREATE TRIGGER IF NOT EXISTS payloads_ad AFTER DELETE ON payloads BEGIN
			DELETE FROM payloads_fts WHERE rowid=old.id;
		END;""")
		c.execute("""CREATE TRIGGER IF NOT EXISTS payloads_au AFTER UPDATE ON payloads BEGIN
			UPDATE payloads_fts SET name=new.name, content=new.content, tags=new.tags
			WHERE rowid=new.id;
		END;""")
		# Encrypted vault
		c.execute("""CREATE TABLE IF NOT EXISTS encrypted_vault (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			title TEXT NOT NULL,
			ciphertext BLOB NOT NULL,
			created_at TEXT, updated_at TEXT
		)""")
		# LainKusanagi OSCP practice list progress tracking
		c.execute("""CREATE TABLE IF NOT EXISTS checklist (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			platform TEXT NOT NULL,
			os TEXT NOT NULL,
			kind TEXT NOT NULL,
			name TEXT NOT NULL,
			status TEXT NOT NULL DEFAULT 'locked',
			notes TEXT DEFAULT '',
			machine_id INTEGER,
			created_at TEXT,
			updated_at TEXT,
			UNIQUE(platform, name)
		)""")
		self.conn.commit()

	def list_notes(self, phase="all"):
		c = self.conn.cursor()
		if phase != "all":
			rows = c.execute(
				"SELECT id,title,phase,tags,updated_at FROM notes WHERE phase=?"
				" ORDER BY title COLLATE NOCASE",
				(phase,),
			).fetchall()
		else:
			rows = c.execute(
				"SELECT id,title,phase,tags,updated_at FROM notes"
				" ORDER BY phase, title COLLATE NOCASE"
			).fetchall()
		return rows

	def get(self, note_id):
		c = self.conn.cursor()
		return c.execute("SELECT * FROM notes WHERE id=?", (note_id,)).fetchone()

	def create(self, title="Untitled", phase="general", tags="", content=""):
		now = datetime.now().isoformat(timespec="seconds")
		c = self.conn.cursor()
		c.execute(
			"INSERT INTO notes(title,content,phase,tags,created_at,updated_at)"
			" VALUES(?,?,?,?,?,?)",
			(title, content, phase, tags, now, now),
		)
		self.conn.commit()
		return c.lastrowid

	def update(self, note_id, **fields):
		fields["updated_at"] = datetime.now().isoformat(timespec="seconds")
		sets = ", ".join(f"{k}=?" for k in fields.keys())
		vals = list(fields.values()) + [note_id]
		self.conn.execute(f"UPDATE notes SET {sets} WHERE id=?", vals)
		self.conn.commit()

	def delete(self, note_id):
		self.conn.execute("DELETE FROM notes WHERE id=?", (note_id,))
		self.conn.commit()

	def search(self, query):
		if not query.strip():
			return self.list_notes()
		tokens = re.findall(r"\w+", query)
		if not tokens:
			return self.list_notes()
		fts_q = " ".join(f'"{t}"' for t in tokens)
		try:
			rows = self.conn.execute(
				"""SELECT n.id,n.title,n.phase,n.tags,n.updated_at
				 FROM notes n JOIN notes_fts f ON f.rowid=n.id
				 WHERE notes_fts MATCH ? ORDER BY rank LIMIT 100""",
				(fts_q,),
			).fetchall()
		except sqlite3.OperationalError:
			return self.list_notes()
		return rows

	def seed_if_empty(self, cheats):
		n = self.conn.execute("SELECT COUNT(*) FROM notes").fetchone()[0]
		if n == 0:
			for c in cheats:
				self.create(
					title=c["title"],
					phase=c["phase"],
					tags=c.get("tags", ""),
					content=c["content"],
				)

	def get_pref(self, key, default=None):
		row = self.conn.execute("SELECT value FROM prefs WHERE key=?", (key,)).fetchone()
		return row["value"] if row else default

	def set_pref(self, key, value):
		self.conn.execute(
			"INSERT INTO prefs(key, value) VALUES(?, ?)"
			" ON CONFLICT(key) DO UPDATE SET value=excluded.value",
			(key, str(value)),
		)
		self.conn.commit()

	# ---------- machine tracker ----------
	def list_machines(self):
		return self.conn.execute(
			"SELECT * FROM machines ORDER BY status='owned' ASC, status='active' DESC, name COLLATE NOCASE"
		).fetchall()

	def get_machine(self, mid):
		return self.conn.execute("SELECT * FROM machines WHERE id=?", (mid,)).fetchone()

	def create_machine(self, name, address="", difficulty="unknown", methodology=None, notes=""):
		now = datetime.now().isoformat(timespec="seconds")
		c = self.conn.cursor()
		c.execute(
			"INSERT INTO machines(name,address,difficulty,started_at,created_at,updated_at,notes)"
			" VALUES(?,?,?,?,?,?,?)",
			(name, address, difficulty, now, now, now, notes),
		)
		mid = c.lastrowid
		# Seed default methodology
		if methodology:
			for phase, items in methodology.items():
				for item in items:
					c.execute(
						"INSERT OR IGNORE INTO machine_checklist(machine_id,phase,item) VALUES(?,?,?)",
						(mid, phase, item),
					)
		self.conn.commit()
		return mid

	def update_machine(self, mid, **fields):
		fields["updated_at"] = datetime.now().isoformat(timespec="seconds")
		sets = ", ".join(f"{k}=?" for k in fields.keys())
		vals = list(fields.values()) + [mid]
		self.conn.execute(f"UPDATE machines SET {sets} WHERE id=?", vals)
		self.conn.commit()

	def delete_machine(self, mid):
		self.conn.execute("DELETE FROM machines WHERE id=?", (mid,))
		self.conn.commit()

	def add_cred(self, machine_id, username, password, secret="", source=""):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"INSERT INTO machine_creds(machine_id,username,password,secret,source,found_at)"
			" VALUES(?,?,?,?,?,?)",
			(machine_id, username, password, secret, source, now),
		)
		self.conn.commit()

	def list_creds(self, machine_id):
		return self.conn.execute(
			"SELECT * FROM machine_creds WHERE machine_id=? ORDER BY id", (machine_id,)
		).fetchall()

	def delete_cred(self, cid):
		self.conn.execute("DELETE FROM machine_creds WHERE id=?", (cid,))
		self.conn.commit()

	def add_loot(self, machine_id, name, kind, value):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"INSERT INTO machine_loot(machine_id,name,kind,value,captured_at) VALUES(?,?,?,?,?)",
			(machine_id, name, kind, value, now),
		)
		self.conn.commit()

	def list_loot(self, machine_id):
		return self.conn.execute(
			"SELECT * FROM machine_loot WHERE machine_id=? ORDER BY id", (machine_id,)
		).fetchall()

	def delete_loot(self, lid):
		self.conn.execute("DELETE FROM machine_loot WHERE id=?", (lid,))
		self.conn.commit()

	def add_screen(self, machine_id, label, path):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"INSERT INTO machine_screens(machine_id,label,path,taken_at) VALUES(?,?,?,?)",
			(machine_id, label, path, now),
		)
		self.conn.commit()

	def list_screens(self, machine_id):
		return self.conn.execute(
			"SELECT * FROM machine_screens WHERE machine_id=? ORDER BY id", (machine_id,)
		).fetchall()

	def delete_screen(self, sid):
		self.conn.execute("DELETE FROM machine_screens WHERE id=?", (sid,))
		self.conn.commit()

	def list_checklist(self, machine_id):
		return self.conn.execute(
			"SELECT * FROM machine_checklist WHERE machine_id=? ORDER BY phase, id",
			(machine_id,),
		).fetchall()

	def tick_checklist(self, item_id, ticked):
		self.conn.execute("UPDATE machine_checklist SET ticked=? WHERE id=?", (1 if ticked else 0, item_id))
		self.conn.commit()

	def checklist_progress(self, machine_id):
		rows = self.list_checklist(machine_id)
		if not rows:
			return 0, 0
		done = sum(1 for r in rows if r["ticked"])
		return done, len(rows)

	# ---------- payloads ----------
	def list_payloads(self, category=None):
		if category != "all":
			return self.conn.execute(
				"SELECT * FROM payloads WHERE category=? ORDER BY name COLLATE NOCASE",
				(category,),
			).fetchall()
		return self.conn.execute(
			"SELECT * FROM payloads ORDER BY category, name COLLATE NOCASE"
		).fetchall()

	def get_payload(self, pid):
		return self.conn.execute("SELECT * FROM payloads WHERE id=?", (pid,)).fetchone()

	def create_payload(self, name, category, platform, content, tags=""):
		now = datetime.now().isoformat(timespec="seconds")
		c = self.conn.cursor()
		c.execute(
			"INSERT INTO payloads(name,category,platform,content,tags,created_at)"
			" VALUES(?,?,?,?,?,?)",
			(name, category, platform, content, tags, now),
		)
		self.conn.commit()
		return c.lastrowid

	def update_payload(self, pid, **fields):
		sets = ", ".join(f"{k}=?" for k in fields.keys())
		vals = list(fields.values()) + [pid]
		self.conn.execute(f"UPDATE payloads SET {sets} WHERE id=?", vals)
		self.conn.commit()

	def delete_payload(self, pid):
		self.conn.execute("DELETE FROM payloads WHERE id=?", (pid,))
		self.conn.commit()

	def search_payloads(self, query):
		if not query.strip():
			return self.list_payloads()
		tokens = re.findall(r"\w+", query)
		if not tokens:
			return self.list_payloads()
		fts_q = " ".join(f'"{t}"' for t in tokens)
		try:
			return self.conn.execute(
				"""SELECT p.* FROM payloads p JOIN payloads_fts f ON f.rowid=p.id
				 WHERE payloads_fts MATCH ? ORDER BY rank LIMIT 100""",
				(fts_q,),
			).fetchall()
		except sqlite3.OperationalError:
			return self.list_payloads()

	def list_payload_categories(self):
		return [r["category"] for r in self.conn.execute(
			"SELECT DISTINCT category FROM payloads WHERE category IS NOT NULL AND category != '' ORDER BY category"
		).fetchall()]

	def seed_payloads_if_empty(self, payload_list):
		n = self.conn.execute("SELECT COUNT(*) FROM payloads").fetchone()[0]
		if n == 0:
			for p in payload_list:
				self.create_payload(
					p["name"], p.get("category", ""), p.get("platform", ""),
					p["content"], p.get("tags", ""),
				)

	# ---------- encrypted vault ----------
	def list_vault(self):
		# Return id, title, created_at, updated_at, length(ciphertext) only
		return self.conn.execute(
			"SELECT id,title,created_at,updated_at,length(ciphertext) AS size FROM encrypted_vault ORDER BY title COLLATE NOCASE"
		).fetchall()

	def get_vault(self, vid):
		return self.conn.execute(
			"SELECT * FROM encrypted_vault WHERE id=?", (vid,)
		).fetchone()

	def create_vault(self, title, ciphertext):
		now = datetime.now().isoformat(timespec="seconds")
		c = self.conn.cursor()
		c.execute(
			"INSERT INTO encrypted_vault(title,ciphertext,created_at,updated_at) VALUES(?,?,?,?)",
			(title, ciphertext, now, now),
		)
		self.conn.commit()
		return c.lastrowid

	def update_vault(self, vid, title, ciphertext):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"UPDATE encrypted_vault SET title=?, ciphertext=?, updated_at=? WHERE id=?",
			(title, ciphertext, now, vid),
		)
		self.conn.commit()

	def delete_vault(self, vid):
		self.conn.execute("DELETE FROM encrypted_vault WHERE id=?", (vid,))
		self.conn.commit()

	# ----------------------------- checklist ----------------------------- #

	def seed_checklist_if_empty(self, rows):
		"""rows: list of (platform, os, kind, name_display). Idempotent — preserves existing progress."""
		now = datetime.now().isoformat(timespec="seconds")
		# Look up what's already in the table for each platform, only insert missing names.
		existing_by_platform = {}
		for plat, *_ in rows:
			if plat in existing_by_platform:
				continue
			existing_by_platform[plat] = {r[0] for r in self.conn.execute(
				"SELECT name FROM checklist WHERE platform=?", (plat,)).fetchall()}
		inserted = 0
		for platform, os_name, kind, name in rows:
			if name in existing_by_platform.get(platform, set()):
				continue
			try:
				self.conn.execute(
					"INSERT INTO checklist(platform,os,kind,name,status,created_at,updated_at)"
					" VALUES (?,?,?,?,?,?,?)",
					(platform, os_name, kind, name, "locked", now, now),
				)
				inserted += 1
			except Exception:
				# UNIQUE collision — already exists
				pass
		if inserted:
			self.conn.commit()

	def list_checklist_items(self, platform=None, status=None, search=None):
		c = self.conn.cursor()
		q = "SELECT id, platform, os, kind, name, status, notes, machine_id FROM checklist WHERE 1=1"
		args = []
		if platform is not None and platform != "all":
			q += " AND platform=?"; args.append(platform)
		if status is not None and status != "all":
			q += " AND status=?"; args.append(status)
		if search:
			q += " AND (name LIKE ? OR notes LIKE ?)"
			args.extend([f"%{search}%", f"%{search}%"])
		q += " ORDER BY platform, os, name COLLATE NOCASE"
		return c.execute(q, args).fetchall()

	def update_checklist_status(self, item_id, status):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"UPDATE checklist SET status=?, updated_at=? WHERE id=?",
			(status, now, item_id))
		self.conn.commit()

	def update_checklist_notes(self, item_id, notes):
		now = datetime.now().isoformat(timespec="seconds")
		self.conn.execute(
			"UPDATE checklist SET notes=?, updated_at=? WHERE id=?",
			(notes, now, item_id))
		self.conn.commit()

	def update_checklist_machine(self, item_id, machine_id):
		self.conn.execute(
			"UPDATE checklist SET machine_id=? WHERE id=?",
			(machine_id, item_id))
		self.conn.commit()

	def checklist_platforms(self):
		return [r[0] for r in self.conn.execute(
			"SELECT DISTINCT platform FROM checklist ORDER BY platform").fetchall()]

	def checklist_stats(self):
		c = self.conn.cursor()
		total = c.execute("SELECT COUNT(*) FROM checklist").fetchone()[0]
		by_status = dict(c.execute(
			"SELECT status, COUNT(*) FROM checklist GROUP BY status").fetchall())
		by_platform = dict(c.execute(
			"SELECT platform, COUNT(*) FROM checklist GROUP BY platform").fetchall())
		return {"total": total, "by_status": by_status, "by_platform": by_platform}

	def close(self):
		self.conn.close()


WELCOME_MD = """# Welcome to OSCP Notes

This is your private, offline note-taking app for the **OSCP exam**.

## Getting started
- Click **File -> New Note** (or press `Cmd+N`) to create a note
- Use the **Phase** dropdown to categorize: `enum`, `exploit`, `privesc`, `post`, `ad`, `cheatsheet`, `general`
- Use **Tags** for free-form labels (comma-separated)
- The right pane shows a live Markdown preview with syntax highlighting
- All your notes are stored locally in `~/OSCP-Notes/data/notes.db` (SQLite)

## Built-in cheat sheets
Open the **CheatSheets** menu to insert ready-made templates:
- Reverse Shells
- Linux / Windows Privilege Escalation
- Active Directory
- File Transfer, Pivoting, Buffer Overflow, Web Enums, One-Liners...

## Markdown supported
- `# H1` to `###### H6`
- **bold**, *italic*
- `inline code`
- Fenced code blocks with language hints:
```bash
nmap -sC -sV -p- TARGET
```
- Lists, blockquotes, tables, hr

## Tips
1. One note per machine - title it like `HTB-Active`, `PG-Play-Buff`
2. Tag every command you actually ran with the right phase
3. Use `Cmd+F` to instantly find any command across all notes
4. `Cmd+E` hides the preview when you want maximum editor space
5. Export any note as `.md` for sharing or backup

Good luck - try harder.
"""


class App:
	def __init__(self, root):
		self.root = root
		self.root.title(f"{APP_NAME} v{APP_VERSION}")
		self.root.geometry("1380x860")
		self.root.minsize(1100, 640)
		self.root.configure(bg=BG)

		self.store = Store(DB_PATH)
		try:
			from cheatseeds import CHEATSHEETS
			self.store.seed_if_empty(CHEATSHEETS)
		except Exception:
			traceback.print_exc()
		try:
			from cheatseeds import PAYLOADS
			self.store.seed_payloads_if_empty(PAYLOADS)
		except Exception:
			traceback.print_exc()
		try:
			import oscp_practice_list as _opl
			rows = [(*m[:3], m[4]) for m in _opl.flat_machines()]
			self.store.seed_checklist_if_empty(rows)
		except Exception:
			traceback.print_exc()

		self.current_id = None
		self.current_phase_filter = "all"
		self._save_after_id = None
		self._preview_after_id = None
		self._dirty = False
		self._suppress_change = False
		self._load_token = 0

		self._configure_styles()
		self._build_menu()
		self._build_layout()
		self._bind_shortcuts()

		# Apply saved theme (must come AFTER layout, _apply_theme touches widgets)
		saved_theme = self.store.get_pref("theme", DEFAULT_THEME)
		if saved_theme in THEMES:
			self._apply_theme(saved_theme, persist=False)

		self._refresh_tree()

		notes = self.store.list_notes()
		if notes:
			self._select_note(notes[0]["id"])
		else:
			nid = self.store.create(title="Welcome", phase="general",
				tags="intro,getting-started", content=WELCOME_MD)
			self._refresh_tree()
			self._select_note(nid)

		self._update_status(f"Loaded {len(notes) or 1} note(s)")

	def _configure_styles(self):
		style = ttk.Style()
		try:
			style.theme_use("clam")
		except tk.TclError:
			pass
		style.configure("Treeview", background=BG_ALT, foreground=FG,
			fieldbackground=BG_ALT, rowheight=26, borderwidth=0,
			font=(UI_FONT, 12))
		style.configure("Treeview.Heading", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11, "bold"))
		style.map("Treeview", background=[("selected", SELECT_BG)],
			foreground=[("selected", SELECT_FG)])
		style.configure("Phase.TButton", background=BG_ALT, foreground=FG,
			borderwidth=0, padding=(10, 6), font=(UI_FONT, 11))
		style.map("Phase.TButton",
			background=[("active", ACCENT_HOVER), ("selected", ACCENT)],
			foreground=[("active", BG), ("selected", BG_ALT)])
		style.configure("TFrame", background=BG)
		style.configure("Status.TLabel", background=BG_ALT, foreground=FG_DIM,
			font=(UI_FONT, 10), padding=(8, 4))

	def _build_menu(self):
		menubar = tk.Menu(self.root)

		m_file = tk.Menu(menubar, tearoff=0)
		m_file.add_command(label="New Note", accelerator=_accel("Cmd+N"), command=self.action_new)
		m_file.add_command(label="Save", accelerator=_accel("Cmd+S"), command=self.action_save)
		m_file.add_command(label="Duplicate", accelerator=_accel("Cmd+D"), command=self.action_duplicate)
		m_file.add_command(label="Delete", accelerator="Delete", command=self.action_delete)
		m_file.add_separator()
		m_file.add_command(label="Export as .md...", command=self.action_export)
		m_file.add_command(label="Import .md...", command=self.action_import)
		m_file.add_separator()
		m_file.add_command(label="Reveal Data Folder", command=self.action_reveal_data)
		m_file.add_separator()
		m_file.add_command(label="Quit", accelerator=_accel("Cmd+Q"), command=self._on_quit)
		menubar.add_cascade(label="File", menu=m_file)

		m_edit = tk.Menu(menubar, tearoff=0)
		m_edit.add_command(label="Find / Search", accelerator=_accel("Cmd+F"), command=self._focus_search)
		m_edit.add_command(label="Toggle Preview", accelerator=_accel("Cmd+E"), command=self._toggle_preview)
		menubar.add_cascade(label="Edit", menu=m_edit)

		m_view = tk.Menu(menubar, tearoff=0)
		m_theme = tk.Menu(m_view, tearoff=0)
		for name in THEMES.keys():
			m_theme.add_command(label=name, command=lambda n=name: self._apply_theme(n))
		m_view.add_cascade(label="Theme", menu=m_theme)
		m_view.add_command(label="Refresh Preview", command=self._refresh_preview)
		menubar.add_cascade(label="View", menu=m_view)

		m_cheat = tk.Menu(menubar, tearoff=0)
		cheats = [
			"Reverse Shells",
			"Linux Privilege Escalation",
			"Windows Privilege Escalation",
			"Active Directory",
			"File Transfer to Target",
			"Web Enumeration",
			"Port Scanning & Service Enum",
			"Pivoting & Tunneling",
			"Buffer Overflow Workflow",
			"Web Shells & Stable Access",
			"Useful One-Liners",
		]
		for name in cheats:
			m_cheat.add_command(label=f"Insert {name}",
				command=lambda n=name: self._insert_cheat(n))
		menubar.add_cascade(label="CheatSheets", menu=m_cheat)

		m_tools = tk.Menu(menubar, tearoff=0)
		m_tools.add_command(label="Open Tracker Tab", command=lambda: self._select_tab("tracker"))
		m_tools.add_command(label="Open Checklist Tab", command=lambda: self._select_tab("checklist"))
		m_tools.add_command(label="Open Payloads Tab", command=lambda: self._select_tab("payloads"))
		m_tools.add_command(label="Open Vault Tab", command=lambda: self._select_tab("vault"))
		m_tools.add_separator()
		m_tools.add_command(label="Export Note as...", command=self.action_export)
		m_tools.add_command(label="Export All Notes as CSV", command=lambda: self.action_bulk_export("csv"))
		m_tools.add_command(label="Export All Notes as JSON", command=lambda: self.action_bulk_export("json"))
		m_tools.add_command(label="Export All Notes as XLSX", command=lambda: self.action_bulk_export("xlsx"))
		m_tools.add_command(label="Export All Notes as PDF", command=lambda: self.action_bulk_export("pdf"))
		m_tools.add_separator()
		m_tools.add_command(label="Generate Report Shell from Note...", command=self.action_generate_report)
		menubar.add_cascade(label="Tools", menu=m_tools)

		m_help = tk.Menu(menubar, tearoff=0)
		m_help.add_command(label="About OSCP Notes", command=self._show_about)
		menubar.add_cascade(label="Help", menu=m_help)

		self.root.config(menu=menubar)

	def _build_layout(self):
		self.status = ttk.Label(self.root, text="", style="Status.TLabel", anchor="w")
		self.status.pack(side="bottom", fill="x")

		top = tk.Frame(self.root, bg=BG, height=44)
		top.pack(side="top", fill="x")
		top.pack_propagate(False)
		ttk.Label(top, text=" Phase:", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11)).pack(side="left", padx=(8, 4))
		self.phase_buttons = {}
		for p in ["all"] + PHASES:
			label = "All" if p == "all" else PHASE_LABELS[p]
			b = ttk.Button(top, text=label, style="Phase.TButton",
				command=lambda pp=p: self._filter_phase(pp))
			b.pack(side="left", padx=2, pady=6)
			self.phase_buttons[p] = b
		ttk.Label(top, text="Search:", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11)).pack(side="right", padx=(8, 4))
		self.search_var = tk.StringVar()
		self.search_var.trace_add("write", lambda *_: self._on_search())
		self.search_entry = ttk.Entry(top, textvariable=self.search_var, width=28,
			font=(UI_FONT, 12))
		self.search_entry.pack(side="right", padx=8, pady=8, ipady=2)



		# Tab strip
		self.notebook = ttk.Notebook(self.root)
		self.notebook.pack(fill="both", expand=True, padx=4, pady=(0, 4))
		self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
		self._tab_frames = {}

		main = tk.PanedWindow(self.root, orient="horizontal", sashwidth=6,
			background=BG, borderwidth=0)
		self.main_paned = main
		# Place main into the Notes tab frame
		notes_tab = tk.Frame(self.notebook, bg=BG)
		self.notebook.add(notes_tab, text="Notes")
		self._tab_frames["notes"] = notes_tab
		main.pack(in_=notes_tab, fill="both", expand=True)

		left = tk.Frame(main, bg=BG_ALT)
		main.add(left, minsize=240, width=280)

		self.tree = ttk.Treeview(left, show="tree", selectmode="browse")
		self.tree.pack(fill="both", expand=True, padx=4, pady=4)
		self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
		self.tree.bind("<Button-3>", self._on_tree_rightclick)
		self.tree.bind("<Button-2>", self._on_tree_rightclick)
		self.tree.bind("<Delete>", lambda e: self.action_delete())

		self.ctx = tk.Menu(self.root, tearoff=0)
		self.ctx.add_command(label="Open", command=self._ctx_open)
		self.ctx.add_command(label="Duplicate", command=self.action_duplicate)
		self.ctx.add_separator()
		for p in PHASES:
			self.ctx.add_command(label=f"Move to {PHASE_LABELS[p]}",
				command=lambda pp=p: self._ctx_move_phase(pp))
		self.ctx.add_separator()
		self.ctx.add_command(label="Export as .md", command=self.action_export)
		self.ctx.add_command(label="Delete", command=self.action_delete)

		center = tk.Frame(main, bg=BG)
		main.add(center, minsize=420)

		hdr = tk.Frame(center, bg=BG)
		hdr.pack(fill="x", padx=12, pady=(10, 4))
		ttk.Label(hdr, text="Title:", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11)).pack(side="left")
		self.title_var = tk.StringVar()
		self.title_entry = ttk.Entry(hdr, textvariable=self.title_var, width=40,
			font=(UI_FONT, 14, "bold"))
		self.title_entry.pack(side="left", padx=8, fill="x", expand=True, ipady=2)
		self.title_var.trace_add("write", lambda *_: self._on_change())

		tags_row = tk.Frame(center, bg=BG)
		tags_row.pack(fill="x", padx=12, pady=(0, 4))
		ttk.Label(tags_row, text="Tags:", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11)).pack(side="left")
		self.tags_var = tk.StringVar()
		self.tags_entry = ttk.Entry(tags_row, textvariable=self.tags_var,
			font=(UI_FONT, 11))
		self.tags_entry.pack(side="left", padx=8, fill="x", expand=True, ipady=2)
		self.tags_var.trace_add("write", lambda *_: self._on_change())

		phase_row = tk.Frame(center, bg=BG)
		phase_row.pack(fill="x", padx=12, pady=(0, 8))
		ttk.Label(phase_row, text="Phase:", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11)).pack(side="left")
		self.phase_var = tk.StringVar(value="general")
		self.phase_combo = ttk.Combobox(phase_row, textvariable=self.phase_var,
			values=PHASES, state="readonly", width=20, font=(UI_FONT, 11))
		self.phase_combo.pack(side="left", padx=8)
		self.phase_combo.bind("<<ComboboxSelected>>", lambda e: self._on_change())

		editor_frame = tk.Frame(center, bg=CODE_BG, highlightthickness=1,
			highlightbackground=BORDER)
		editor_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
		self.editor = tk.Text(editor_frame, bg=BG, fg=FG, insertbackground=FG,
			selectbackground=SELECT_BG, selectforeground=SELECT_FG,
			font=(MONO_FONT, 13), wrap="word", undo=True, relief="flat",
			padx=10, pady=10, highlightthickness=0)
		self.editor.pack(fill="both", expand=True)
		self.editor.tag_configure("codeblock", background=CODE_BG, foreground=CODE_FG,
			font=(MONO_FONT, 12), lmargin1=10, lmargin2=10)
		self.editor.tag_configure("code_line", background=CODE_BG, foreground=CODE_FG)
		self.editor.tag_configure("heading", foreground=THEMES[_ACTIVE_THEME]["H1"],
			font=(MONO_FONT, 14, "bold"))
		self.editor.bind("<<Modified>>", self._on_editor_modified)
		self.editor.bind("<KeyRelease>", lambda e: self._on_change())

		right = tk.Frame(main, bg=BG)
		main.add(right, minsize=320, width=480)
		self.preview_frame = right
		ttk.Label(right, text="Preview", background=BG, foreground=FG_DIM,
			font=(UI_FONT, 11, "bold")).pack(anchor="w", padx=12, pady=(10, 0))

		preview_inner = tk.Frame(right, bg=BG, highlightthickness=1,
			highlightbackground=BORDER)
		preview_inner.pack(fill="both", expand=True, padx=12, pady=(4, 12))
		self.preview = tk.Text(preview_inner, bg=BG, fg=FG,
			font=(UI_FONT, 12), wrap="word", relief="flat",
			padx=14, pady=10, highlightthickness=0, state="disabled", cursor="arrow")
		self.preview.pack(fill="both", expand=True)
		self.preview.tag_configure("h1", font=(UI_FONT, 22, "bold"),
			foreground=THEMES[_ACTIVE_THEME]["H1"], spacing1=8, spacing3=6)
		self.preview.tag_configure("h2", font=(UI_FONT, 17, "bold"),
			foreground=THEMES[_ACTIVE_THEME]["H2"], spacing1=6, spacing3=4)
		self.preview.tag_configure("h3", font=(UI_FONT, 14, "bold"),
			foreground=THEMES[_ACTIVE_THEME]["H3"], spacing1=4, spacing3=2)
		self.preview.tag_configure("h4", font=(UI_FONT, 13, "bold"),
			foreground=THEMES[_ACTIVE_THEME]["H4"])
		self.preview.tag_configure("p", spacing1=2, spacing3=4)
		self.preview.tag_configure("li", lmargin1=14, lmargin2=28)
		self.preview.tag_configure("code", background=CODE_BG, foreground=CODE_FG,
			font=(MONO_FONT, 12))
		self.preview.tag_configure("codeblock", background=CODE_BG, foreground=CODE_FG,
			font=(MONO_FONT, 12), lmargin1=10, lmargin2=10, spacing1=4, spacing3=4)
		self.preview.tag_configure("quote", foreground=THEMES[_ACTIVE_THEME]["QUOTE"], lmargin1=14,
			lmargin2=14, font=(UI_FONT, 12, "italic"))
		self.preview.tag_configure("hr", foreground=BORDER)
		self.preview.tag_configure("bold", font=(UI_FONT, 12, "bold"))
		self.preview.tag_configure("italic", font=(UI_FONT, 12, "italic"))
		self.preview.tag_configure("a", foreground=THEMES[_ACTIVE_THEME]["LINK"], underline=True)

	def _bind_shortcuts(self):
		# Cross-platform: <Command-*> on macOS, <Control-*> on Windows/Linux.
		# Tk only recognizes the <Command> modifier on macOS, so we have to
		# bind both forms everywhere to keep the app consistent.
		shortcuts = [
			("n", self.action_new), ("s", self.action_save),
			("d", self.action_duplicate), ("f", self._focus_search),
			("e", self._toggle_preview), ("q", self._on_quit),
		]
		for key, fn in shortcuts:
			for modifier in ("Command", "Control"):
				for case in (key, key.upper()):
					self.root.bind(f"<{modifier}-{case}>", lambda e, fn=fn: fn())
		self.root.bind("<Delete>", lambda e: self.action_delete())
		self.root.protocol("WM_DELETE_WINDOW", self._on_quit)

	def _refresh_tree(self, rows=None):
		self.tree.delete(*self.tree.get_children())
		if rows is None:
			if self.current_phase_filter == "all":
				rows = self.store.list_notes()
			else:
				rows = self.store.list_notes(self.current_phase_filter)

		groups = {p: [] for p in PHASES}
		for r in rows:
			groups.setdefault(r["phase"], []).append(r)

		for p in PHASES:
			if not groups[p]:
				continue
			label = f"{PHASE_LABELS[p]} ({len(groups[p])})"
			node = self.tree.insert("", "end", iid=f"phase:{p}", text=label, open=True)
			for n in groups[p]:
				self.tree.insert(node, "end", iid=f"note:{n['id']}",
					text=f" {n['title']} ", values=(n["id"],))

	def _filter_phase(self, phase):
		self.current_phase_filter = phase
		self.search_var.set("")
		self._refresh_tree()

	def _on_tree_select(self, _evt=None):
		sel = self.tree.selection()
		if not sel:
			return
		tag = sel[0]
		if tag.startswith("note:"):
			nid = int(tag.split(":", 1)[1])
			self._select_note(nid)

	def _on_tree_rightclick(self, event):
		rowid = self.tree.identify_row(event.y)
		if not rowid:
			return
		self.tree.selection_set(rowid)
		if rowid.startswith("note:"):
			try:
				self.ctx.tk_popup(event.x_root, event.y_root)
			finally:
				self.ctx.grab_release()

	def _ctx_open(self):
		sel = self.tree.selection()
		if sel[0].startswith("note:"):
			self._select_note(int(sel[0].split(":", 1)[1]))

	def _ctx_move_phase(self, phase):
		if self.current_id is None:
			return
		self.store.update(self.current_id, phase=phase)
		self.phase_var.set(phase)
		self._refresh_tree()
		self._flash_status(f"Moved to {PHASE_LABELS[phase]}")

	def _on_search(self):
		q = self.search_var.get()
		if not q.strip():
			self._refresh_tree()
			return
		rows = self.store.search(q)
		self.tree.delete(*self.tree.get_children())
		node = self.tree.insert("", "end", iid="phase:results",
			text=f"Results ({len(rows)})", open=True)
		for r in rows:
			self.tree.insert(node, "end", iid=f"note:{r['id']}",
				text=f" {r['title']} ", values=(r["id"],))

	def _focus_search(self):
		self.search_entry.focus_set()
		self.search_entry.select_range(0, "end")

	def _toggle_preview(self):
		if self.preview_frame.winfo_ismapped():
			self.main_paned.forget(self.preview_frame)
			self._flash_status("Preview hidden")
		else:
			self.main_paned.add(self.preview_frame, minsize=320, width=480)
			self._refresh_preview()
			self._flash_status("Preview shown")

	def _select_note(self, note_id):
		if self._dirty and self.current_id is not None:
			self._do_save()
		if self.current_id == note_id:
			return
		row = self.store.get(note_id)
		if not row:
			return
		self.current_id = note_id
		self._suppress_change = True
		self.title_var.set(row["title"])
		self.tags_var.set(row["tags"])
		self.phase_var.set(row["phase"])
		self.editor.delete("1.0", "end")
		self.editor.insert("1.0", row["content"])
		self._suppress_change = False
		self._dirty = False
		self.editor.edit_reset()
		self.editor.edit_modified(False)
		self._apply_editor_syntax()
		self._refresh_preview()
		self._update_status(f"Loaded note #{note_id} - {row['title']} - {row['phase']}")
		iid = f"note:{note_id}"
		if iid in self.tree.get_children():
			self.tree.selection_set(iid)
			self.tree.see(iid)

	def action_new(self):
		title = simpledialog.askstring("New Note", "Title:", initialvalue="New Note",
			parent=self.root)
		if not title:
			return
		phase = self.current_phase_filter if self.current_phase_filter in PHASES else "general"
		nid = self.store.create(title=title, phase=phase, tags="", content=f"# {title}\n\n")
		self._refresh_tree()
		self._select_note(nid)
		self.title_entry.focus_set()
		self.title_entry.select_range(0, "end")

	def action_save(self):
		if self.current_id is None:
			return
		self._do_save()
		self._flash_status("Saved")

	def action_duplicate(self):
		if self.current_id is None:
			return
		row = self.store.get(self.current_id)
		if not row:
			return
		new_id = self.store.create(
			title=row["title"] + " (copy)",
			content=row["content"],
			phase=row["phase"],
			tags=row["tags"],
		)
		self._refresh_tree()
		self._select_note(new_id)
		self._flash_status("Duplicated")

	def action_delete(self):
		if self.current_id is None:
			return
		row = self.store.get(self.current_id)
		if not row:
			return
		if not messagebox.askyesno("Delete note",
			f"Delete '{row['title']}'? This cannot be undone.", parent=self.root):
			return
		self.store.delete(self.current_id)
		deleted_id = self.current_id
		self.current_id = None
		self._refresh_tree()
		rows = self.store.list_notes(self.current_phase_filter)
		if rows:
			self._select_note(rows[0]["id"])
		else:
			nid = self.store.create(title="New Note", phase="general",
				tags="", content="# New Note\n\n")
			self._refresh_tree()
			self._select_note(nid)
		self._flash_status(f"Deleted #{deleted_id}")

	def action_export(self):
		if self.current_id is None:
			messagebox.showinfo("No note selected",
				"Open a note from the Notes tab first (click on it in the left list).",
				parent=self.root)
			return
		row = self.store.get(self.current_id)
		if not row:
			return
		fmt = self._ask_export_format()
		if not fmt:
			return
		ext = {"md": "md", "txt": "txt", "html": "html", "json": "json",
			"pdf": "pdf", "csv": "csv", "xlsx": "xlsx"}.get(fmt, fmt)
		default = (APP_DIR / f"{re.sub(r'[^A-Za-z0-9_-]+', '_', row['title'])}.{ext}").as_posix()
		filetypes = {
			"md": [("Markdown", "*.md")], "txt": [("Text", "*.txt")],
			"html": [("HTML", "*.html")], "json": [("JSON", "*.json")],
			"pdf": [("PDF", "*.pdf")], "csv": [("CSV", "*.csv")],
			"xlsx": [("Excel", "*.xlsx")],
		}.get(fmt, [("All", "*.*")])
		path = filedialog.asksaveasfilename(
			title=f"Export note as {fmt.upper()}", defaultextension=f".{ext}",
			initialfile=Path(default).name, initialdir=APP_DIR,
			filetypes=filetypes, parent=self.root)
		if not path:
			return
		try:
			if fmt == "md":
				Path(path).write_bytes(export_as_markdown(row))
			elif fmt == "txt":
				Path(path).write_bytes(export_as_text(row))
			elif fmt == "html":
				css = "body{font-family:Helvetica,Arial,sans-serif;max-width:900px;margin:2em auto;padding:0 1em;background:#fafafa;color:#222}"
				css += "h1,h2{color:#0e7490} pre{background:#0d0d0d;color:#ce9178;padding:1em;border-radius:6px;overflow:auto}"
				Path(path).write_bytes(export_as_html(row, css))
			elif fmt == "json":
				Path(path).write_bytes(export_as_json(row))
			elif fmt == "csv":
				Path(path).write_bytes(export_as_csv([row]))
			elif fmt == "xlsx":
				export_as_xlsx([row], path)
			elif fmt == "pdf":
				export_as_pdf([row], path, title=f"{row['title']} - OSCP Notes Export")
		except Exception as e:
			messagebox.showerror("Export failed", str(e), parent=self.root)
			return
		self._flash_status(f"Exported {fmt.upper()} -> {Path(path).name}")
		try:
			webbrowser.open(Path(path).parent.as_uri())
		except Exception:
			pass

	def _ask_export_format(self):
		"""Popup to choose export format. Returns 'md', 'txt', 'html', 'json', 'pdf', 'csv', 'xlsx' or None."""
		dlg = tk.Toplevel(self.root)
		dlg.title("Export format")
		dlg.transient(self.root)
		dlg.resizable(False, False)
		dlg.configure(bg=BG)
		dlg.lift()
		dlg.focus_force()
		dlg.grab_set()
		choice = {"value": None}
		ttk.Label(dlg, text="Choose export format:", background=BG, foreground=FG,
			font=(UI_FONT, 11)).grid(row=0, column=0, columnspan=4, padx=12, pady=(12, 8), sticky="w")
		formats = [
			("md", "Markdown"), ("txt", "Plain Text"),
			("html", "HTML"), ("json", "JSON"),
			("csv", "CSV"), ("xlsx", "Excel (xlsx)"),
			("pdf", "PDF"),
		]
		def pick(f):
			choice["value"] = f
			dlg.destroy()
		# arrange as a grid of small buttons
		for i, (f, label) in enumerate(formats):
			r, c = divmod(i, 4)
			b = ttk.Button(dlg, text=label, command=lambda ff=f: pick(ff))
			b.grid(row=1+r, column=c, padx=6, pady=6, sticky="ew")
		ttk.Button(dlg, text="Cancel", command=dlg.destroy).grid(row=3, column=0, columnspan=4, pady=(6, 12))
		self.root.wait_window(dlg)
		return choice["value"]

	def action_bulk_export(self, fmt: str):
		notes = self.store.list_notes()
		if not notes:
			messagebox.showinfo("Nothing to export", "No notes yet.", parent=self.root)
			return
		stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
		ext = {"csv": "csv", "json": "json", "xlsx": "xlsx", "pdf": "pdf"}[fmt]
		default = (APP_DIR / f"oscp_notes_{stamp}.{ext}").as_posix()
		path = filedialog.asksaveasfilename(
			title=f"Export ALL notes as {fmt.upper()}", defaultextension=f".{ext}",
			initialfile=Path(default).name, initialdir=APP_DIR,
			filetypes=[(fmt.upper(), f"*.{ext}")], parent=self.root)
		if not path:
			return
		try:
			if fmt == "csv":
				Path(path).write_bytes(export_as_csv(notes))
			elif fmt == "json":
				Path(path).write_text(
					json.dumps([dict(n) for n in notes], indent=2, ensure_ascii=False),
					encoding="utf-8",
				)
			elif fmt == "xlsx":
				export_as_xlsx(notes, path)
			elif fmt == "pdf":
				export_as_pdf(notes, path, title=f"OSCP Notes - {len(notes)} notes - {stamp}")
		except Exception as e:
			messagebox.showerror("Bulk export failed", str(e), parent=self.root)
			return
		self._flash_status(f"Exported {len(notes)} notes -> {Path(path).name}")
		try:
			webbrowser.open(Path(path).parent.as_uri())
		except Exception:
			pass

	def action_generate_report(self):
		if self.current_id is None:
			messagebox.showinfo("No note", "Open a note first.", parent=self.root)
			return
		row = self.store.get(self.current_id)
		if not row:
			return
		md = generate_report_shell(row)
		new_id = self.store.create(
			title=f"{row['title']} - Report Draft",
			content=md,
			phase="post",
			tags="report-draft",
		)
		# Switch to Notes tab and select the new note
		self._select_tab("notes")
		self._refresh_tree()
		self._select_note(new_id)
		self._flash_status(f"Report shell generated (#{new_id})")

	def action_import(self):
		path = filedialog.askopenfilename(title="Import .md",
			filetypes=[("Markdown", "*.md"), ("All", "*.*")], parent=self.root)
		if not path:
			return
		text = Path(path).read_text(encoding="utf-8", errors="ignore")
		m = re.search(r"^#\s+(.+)$", text, re.M)
		title = m.group(1).strip() if m else Path(path).stem
		body = re.sub(r"^#\s+.+\n+", "", text, count=1)
		nid = self.store.create(title=title, phase="general", tags="imported", content=body)
		self._refresh_tree()
		self._select_note(nid)
		self._flash_status(f"Imported '{title}'")

	def action_reveal_data(self):
		try:
			webbrowser.open(APP_DIR.as_uri())
		except Exception as e:
			messagebox.showerror("Error", str(e))

	def _insert_cheat(self, title):
		row = None
		for r in self.store.list_notes("cheatsheet"):
			if r["title"] == title:
				row = r
				break
		if not row:
			self._flash_status(f"Cheat '{title}' not found")
			return
		new_id = self.store.create(
			title=f"{title} (mine)",
			content=self.store.get(row["id"])["content"],
			phase="general",
			tags="template",
		)
		self._refresh_tree()
		self._select_note(new_id)
		self._flash_status(f"Inserted '{title}' template")

	def _on_editor_modified(self, _evt=None):
		if self._suppress_change:
			self.editor.edit_modified(False)
			return
		self._apply_editor_syntax()
		self.editor.edit_modified(False)

	def _on_change(self):
		if self._suppress_change:
			return
		if self.current_id is None:
			return
		self._dirty = True
		if self._save_after_id:
			self.root.after_cancel(self._save_after_id)
		self._save_after_id = self.root.after(1200, self._do_save)
		if self._preview_after_id:
			self.root.after_cancel(self._preview_after_id)
		self._preview_after_id = self.root.after(400, self._refresh_preview)
		self._update_status("Editing...")

	def _do_save(self):
		if not self._dirty or self.current_id is None:
			return
		self._save_after_id = None
		title = self.title_var.get().strip() or "Untitled"
		self.store.update(
			self.current_id,
			title=title,
			content=self.editor.get("1.0", "end-1c"),
			phase=self.phase_var.get() or "general",
			tags=self.tags_var.get().strip(),
		)
		self._dirty = False
		self._refresh_tree_node(self.current_id, title)
		self._update_status(f"Saved - {datetime.now().strftime('%H:%M:%S')}")

	def _refresh_tree_node(self, note_id, title):
		iid = f"note:{note_id}"
		if iid in self.tree.get_children():
			self.tree.item(iid, text=f" {title} ")

	def _apply_editor_syntax(self):
		if self._suppress_change:
			return
		content = self.editor.get("1.0", "end-1c")
		for tag in ("codeblock", "code_line", "heading"):
			self.editor.tag_remove(tag, "1.0", "end")
		for m in re.finditer(r"^```.*$", content, re.MULTILINE):
			start = f"1.0+{m.start()}c"
			end = f"1.0+{m.end()}c"
			self.editor.tag_add("code_line", start, end)
		for m in re.finditer(r"^#{1,6}\s.*$", content, re.MULTILINE):
			start = f"1.0+{m.start()}c"
			end = f"1.0+{m.end()}c"
			self.editor.tag_add("heading", start, end)
		for m in re.finditer(r"\*\*([^*]+)\*\*", content):
			start = f"1.0+{m.start(1)}c"
			end = f"1.0+{m.end(1)}c"
			self.editor.tag_add("bold", start, end)
		for m in re.finditer(r"(?<!\*)\*([^*\n]+)\*(?!\*)", content):
			start = f"1.0+{m.start(1)}c"
			end = f"1.0+{m.end(1)}c"
			self.editor.tag_add("italic", start, end)

	def _refresh_preview(self):
		if self.current_id is None:
			return
		self._preview_after_id = None
		self._load_token += 1
		token = self._load_token
		md_text = self.editor.get("1.0", "end-1c")
		try:
			html = render_markdown(md_text)
			self._render_preview_html(html, token)
		except Exception as e:
			self.preview.configure(state="normal")
			self.preview.delete("1.0", "end")
			self.preview.insert("1.0", f"Preview error: {e}\n{traceback.format_exc()}")
			self.preview.configure(state="disabled")

	def _render_preview_html(self, html, token):
		if token != self._load_token:
			return
		self.preview.configure(state="normal")
		self.preview.delete("1.0", "end")

		html = re.sub(r"<style.*?</style>", "", html, flags=re.DOTALL | re.IGNORECASE)

		tag_stack = []
		pattern = re.compile(r"<(/?)(\w+)([^>]*)>|([^<]+)", re.DOTALL)
		pos = 0
		for m in pattern.finditer(html):
			if m.start() != pos:
				pass
			pos = m.end()
			if m.group(1) is not None and m.group(1) != "":
				tag = m.group(2).lower()
				if tag_stack[-1] == tag:
					tag_stack.pop()
				elif tag in tag_stack:
					while tag_stack[-1] != tag:
						tag_stack.pop()
					if tag_stack:
						tag_stack.pop()
			elif m.group(2):
				tag = m.group(2).lower()
				if tag == "br":
					self.preview.insert("end", "\n", ())
				elif tag == "hr":
					self.preview.insert("end", "\n" + "-" * 60 + "\n", "hr")
				else:
					tag_stack.append(tag)
			else:
				text = m.group(4)
				if not text:
					continue
				text = (text.replace("&lt;", "<").replace("&gt;", ">")
					.replace("&amp;", "&").replace("&quot;", '"')
					.replace("&#39;", "'").replace("&nbsp;", " "))
				eff = set(tag_stack)
				tags = ()
				if "h1" in eff:
					tags = ("h1",) + tags
				if "h2" in eff:
					tags = ("h2",) + tags
				if "h3" in eff:
					tags = ("h3",) + tags
				if "h4" in eff:
					tags = ("h4",) + tags
				if "p" in eff and "li" not in eff:
					tags = ("p",) + tags
				if "li" in eff:
					tags = ("li",) + tags
				if "blockquote" in eff:
					tags = ("quote",) + tags
				if "strong" in eff or "b" in eff:
					tags = tags + ("bold",)
				if "em" in eff or "i" in eff:
					tags = tags + ("italic",)
				if "code" in eff:
					if "pre" in eff:
						tags = tags + ("codeblock",)
					else:
						tags = tags + ("code",)
				if "a" in eff:
					tags = tags + ("a",)
				self.preview.insert("end", text, tags)

		self.preview.configure(state="disabled")

	def _update_status(self, msg):
		self.status.config(text=f" {msg}")

	def _flash_status(self, msg):
		self._update_status(msg)

	def _apply_theme(self, name, persist=True):
		"""Switch the live palette to a named theme and re-style everything."""
		global _ACTIVE_THEME, BG, BG_ALT, FG, FG_DIM, ACCENT, ACCENT_HOVER
		global SELECT_BG, SELECT_FG, CODE_BG, CODE_FG, BORDER
		if name not in THEMES:
			return
		_ACTIVE_THEME = name
		t = THEMES[name]
		BG = t["BG"]; BG_ALT = t["BG_ALT"]
		FG = t["FG"]; FG_DIM = t["FG_DIM"]
		ACCENT = t["ACCENT"]; ACCENT_HOVER = t["ACCENT_HOVER"]
		SELECT_BG = t["SELECT_BG"]; SELECT_FG = t["SELECT_FG"]
		CODE_BG = t["CODE_BG"]; CODE_FG = t["CODE_FG"]
		BORDER = t["BORDER"]

		# Window + status
		try:
			self.root.configure(bg=BG)
		except Exception:
			pass
		try:
			self.status.configure(background=BG_ALT, foreground=FG_DIM)
		except Exception:
			pass

		# Reconfigure ttk styles
		style = ttk.Style()
		style.configure("Treeview", background=BG_ALT, foreground=FG,
			fieldbackground=BG_ALT)
		style.configure("Treeview.Heading", background=BG, foreground=FG_DIM)
		style.map("Treeview", background=[("selected", SELECT_BG)],
			foreground=[("selected", SELECT_FG)])
		style.configure("Phase.TButton", background=BG_ALT, foreground=FG)
		style.map("Phase.TButton",
			background=[("active", ACCENT_HOVER), ("selected", ACCENT)],
			foreground=[("active", BG), ("selected", BG_ALT)])
		style.configure("Status.TLabel", background=BG_ALT, foreground=FG_DIM)

		# Walk every child widget and re-apply colors based on its role
		def restyle(w):
			cls = w.winfo_class()
			try:
				if cls in ("Frame",):
					# Apply BG to plain frames; BG_ALT to sidebars (heuristic: small width)
					w.configure(bg=BG)
				elif cls == "Label":
					if str(w.cget("text")).startswith(" Phase:") or str(w.cget("text")).startswith("Tags:") \
						or str(w.cget("text")).startswith("Title:") or str(w.cget("text")).startswith("Search:") \
						or str(w.cget("text")) == "Preview":
						w.configure(background=BG, foreground=FG_DIM)
					else:
						w.configure(background=BG, foreground=FG_DIM)
				elif cls == "Entry":
					w.configure(background=BG_ALT, foreground=FG,
						insertbackground=FG, highlightbackground=BORDER,
						highlightcolor=BORDER)
				elif cls == "Text":
					# Editor and preview are distinguished by state
					state = str(w.cget("state"))
					if state == "disabled":
						w.configure(background=BG, foreground=FG)
					else:
						w.configure(background=BG, foreground=FG,
							insertbackground=FG, selectbackground=SELECT_BG,
							selectforeground=SELECT_FG,
							highlightbackground=BORDER)
				elif cls == "TCombobox":
					style.configure("TCombobox",
						fieldbackground=BG_ALT, background=BG_ALT,
						foreground=FG, arrowcolor=FG)
			except Exception:
				pass
			for c in w.winfo_children():
				restyle(c)

		restyle(self.root)

		# Re-tag editor and preview with the new theme
		try:
			self.editor.tag_configure("codeblock", background=CODE_BG, foreground=CODE_FG)
			self.editor.tag_configure("code_line", background=CODE_BG, foreground=CODE_FG)
			self.editor.tag_configure("heading", foreground=THEMES[_ACTIVE_THEME]["H1"])
		except Exception:
			pass
		try:
			self.preview.tag_configure("h1", foreground=THEMES[_ACTIVE_THEME]["H1"])
			self.preview.tag_configure("h2", foreground=THEMES[_ACTIVE_THEME]["H2"])
			self.preview.tag_configure("h3", foreground=THEMES[_ACTIVE_THEME]["H3"])
			self.preview.tag_configure("h4", foreground=THEMES[_ACTIVE_THEME]["H4"])
			self.preview.tag_configure("code", background=CODE_BG, foreground=CODE_FG)
			self.preview.tag_configure("codeblock", background=CODE_BG, foreground=CODE_FG)
			self.preview.tag_configure("quote", foreground=THEMES[_ACTIVE_THEME]["QUOTE"])
			self.preview.tag_configure("hr", foreground=BORDER)
			self.preview.tag_configure("a", foreground=THEMES[_ACTIVE_THEME]["LINK"])
		except Exception:
			pass

		# Force the preview to re-render so syntax colors catch up
		try:
			self._refresh_preview()
		except Exception:
			pass
		try:
			self.root.update_idletasks()
		except Exception:
			pass
		if persist:
			try:
				self.store.set_pref("theme", name)
			except Exception:
				pass
		self._flash_status(f"Theme: {name}")

	def _show_about(self):
		messagebox.showinfo(APP_NAME,
			f"{APP_NAME} v{APP_VERSION}\n\n"
			f"Native macOS note-taking app for OSCP exam preparation.\n\n"
			f"Author: {APP_AUTHOR}\n"
			f"LinkedIn: {APP_LINKEDIN}\n\n"
			f"Data folder: {APP_DIR}\n"
			f"DB: {DB_PATH}\n\n"
			f"Themes (View -> Theme):\n"
			+ "\n".join(f" - {n}" for n in THEMES.keys())
			+ "\n\n"
			f"Shortcuts:\n"
			f" Cmd+N new note\n"
			f" Cmd+S save\n"
			f" Cmd+D duplicate\n"
			f" Cmd+F search\n"
			f" Cmd+E toggle preview\n"
			f" Delete delete note\n"
			f" Cmd+Q quit",
			parent=self.root)

	def _on_quit(self):
		if self._dirty and self.current_id is not None:
			self._do_save()
		self.store.close()
		self.root.destroy()

	# ----------------------------- tab management ----------------------------- #

	def _select_tab(self, name: str):
		# Look up the actual tab by name in the live notebook tabs.
		# This works whether tabs have been built lazily or not.
		tab_text_map = {"notes": "Notes", "tracker": "Tracker", "checklist": "Checklist",
			"payloads": "Payloads", "vault": "Vault"}
		target = tab_text_map.get(name, name)
		for idx in range(self.notebook.index("end")):
			if self.notebook.tab(idx, "text") == target:
				self.notebook.select(idx)
				return
		# Tab not built yet — try building it and selecting again
		if name == "tracker" and not getattr(self, "_tracker_built", False):
			self._build_tracker_tab()
		elif name == "checklist" and not getattr(self, "_checklist_built", False):
			self._build_checklist_tab()
		elif name == "payloads" and not getattr(self, "_payloads_built", False):
			self._build_payloads_tab()
		elif name == "vault" and not getattr(self, "_vault_built", False):
			self._build_vault_tab()
		# Re-look
		for idx in range(self.notebook.index("end")):
			if self.notebook.tab(idx, "text") == target:
				self.notebook.select(idx)
				return

	def _on_tab_changed(self, event):
		# Lazily build tabs the first time they're selected.
		# The user is moving to whichever tab is now selected.
		try:
			current = self.notebook.index(self.notebook.select())
		except Exception:
			return
		try:
			tab_text = self.notebook.tab(current, "text")
		except Exception:
			tab_text = ""
		if tab_text == "Tracker" and not getattr(self, "_tracker_built", False):
			self._build_tracker_tab()
		elif tab_text == "Checklist" and not getattr(self, "_checklist_built", False):
			self._build_checklist_tab()
		elif tab_text == "Payloads" and not getattr(self, "_payloads_built", False):
			self._build_payloads_tab()
		elif tab_text == "Vault" and not getattr(self, "_vault_built", False):
			self._build_vault_tab()

	# ----------------------------- tracker tab ----------------------------- #

	def _build_tracker_tab(self):
		frame = tk.Frame(self.notebook, bg=BG)
		self.notebook.add(frame, text="Tracker")
		self._tab_frames["tracker"] = frame
		self._tracker_built = True

		# Two-pane: machines list on left, details on right
		left = tk.Frame(frame, bg=BG_ALT, width=260)
		left.pack(side="left", fill="y", padx=(4, 2), pady=4)
		left.pack_propagate(False)
		right = tk.Frame(frame, bg=BG)
		right.pack(side="left", fill="both", expand=True, padx=(2, 4), pady=4)

		# Left: machine list
		ttk.Label(left, text="Machines", style="Section.TLabel").pack(anchor="w", padx=8, pady=(8, 4))
		btn_row = tk.Frame(left, bg=BG_ALT)
		btn_row.pack(fill="x", padx=4, pady=2)
		ttk.Button(btn_row, text="+ New", command=self._tracker_new_machine).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Refresh", command=self._tracker_refresh).pack(side="left", padx=2)

		self.machine_list = tk.Listbox(left, bg=BG, fg=FG, selectbackground=SELECT_BG,
			selectforeground=SELECT_FG, font=(UI_FONT, 11), relief="flat", highlightthickness=0)
		self.machine_list.pack(fill="both", expand=True, padx=4, pady=4)
		self.machine_list.bind("<<ListboxSelect>>", self._tracker_on_select)

		# Right: details
		self.tracker_right = right
		self.tracker_machine_id = None
		self._build_tracker_detail(right)
		self._tracker_refresh()

	def _build_tracker_detail(self, parent):
		# Header: name, address, difficulty, status
		hdr = tk.Frame(parent, bg=BG)
		hdr.pack(fill="x", padx=8, pady=(8, 4))

		ttk.Label(hdr, text="Name:", background=BG, foreground=FG_DIM).grid(row=0, column=0, sticky="e", padx=4, pady=2)
		self.m_name_var = tk.StringVar()
		ttk.Entry(hdr, textvariable=self.m_name_var, width=30, font=(UI_FONT, 12, "bold")).grid(row=0, column=1, sticky="ew", padx=4)

		ttk.Label(hdr, text="Address:", background=BG, foreground=FG_DIM).grid(row=0, column=2, sticky="e", padx=4)
		self.m_addr_var = tk.StringVar()
		ttk.Entry(hdr, textvariable=self.m_addr_var, width=20, font=(UI_FONT, 11)).grid(row=0, column=3, sticky="ew", padx=4)

		ttk.Label(hdr, text="Difficulty:", background=BG, foreground=FG_DIM).grid(row=0, column=4, sticky="e", padx=4)
		self.m_diff_var = tk.StringVar(value="unknown")
		diffs = ["easy", "medium", "hard", "insane", "unknown"]
		ttk.Combobox(hdr, textvariable=self.m_diff_var, values=diffs, state="readonly", width=10).grid(row=0, column=5, padx=4)

		ttk.Label(hdr, text="Status:", background=BG, foreground=FG_DIM).grid(row=1, column=0, sticky="e", padx=4, pady=2)
		self.m_status_var = tk.StringVar(value="active")
		ttk.Combobox(hdr, textvariable=self.m_status_var, values=["active", "paused", "owned", "retired"], state="readonly", width=10).grid(row=1, column=1, sticky="w", padx=4)

		ttk.Label(hdr, text="Elapsed:", background=BG, foreground=FG_DIM).grid(row=1, column=2, sticky="e", padx=4)
		self.m_elapsed_var = tk.StringVar(value="--:--:--")
		ttk.Label(hdr, textvariable=self.m_elapsed_var, background=BG, foreground=ACCENT, font=(MONO_FONT, 12, "bold")).grid(row=1, column=3, sticky="w")

		ctl_row = tk.Frame(hdr, bg=BG)
		ctl_row.grid(row=1, column=4, columnspan=2, sticky="e", padx=4)
		ttk.Button(ctl_row, text="Save", command=self._tracker_save_machine).pack(side="left", padx=2)
		ttk.Button(ctl_row, text="▶/⏸", command=self._tracker_toggle_pause, width=4).pack(side="left", padx=2)
		ttk.Button(ctl_row, text="Owned", command=lambda: self._tracker_set_status("owned")).pack(side="left", padx=2)
		ttk.Button(ctl_row, text="Delete", command=self._tracker_delete_machine).pack(side="left", padx=2)

		hdr.columnconfigure(1, weight=1)
		hdr.columnconfigure(3, weight=1)

		# Sub-notebook: creds / loot / screens / checklist / flags / notes
		sub = ttk.Notebook(parent)
		sub.pack(fill="both", expand=True, padx=8, pady=(4, 8))
		self.tracker_sub = sub

		# Creds tab
		creds_f = tk.Frame(sub, bg=BG)
		sub.add(creds_f, text="Credentials")
		self._build_tracker_creds(creds_f)
		# Loot tab
		loot_f = tk.Frame(sub, bg=BG)
		sub.add(loot_f, text="Loot")
		self._build_tracker_loot(loot_f)
		# Screenshots tab
		scr_f = tk.Frame(sub, bg=BG)
		sub.add(scr_f, text="Screenshots")
		self._build_tracker_screens(scr_f)
		# Checklist tab
		chk_f = tk.Frame(sub, bg=BG)
		sub.add(chk_f, text="Checklist")
		self._build_tracker_checklist(chk_f)
		# Flags tab
		flg_f = tk.Frame(sub, bg=BG)
		sub.add(flg_f, text="Flags")
		self._build_tracker_flags(flg_f)
		# Notes tab
		note_f = tk.Frame(sub, bg=BG)
		sub.add(note_f, text="Notes")
		self._build_tracker_notes(note_f)

		# Start the timer tick
		self._schedule_timer_tick()

	def _build_tracker_creds(self, parent):
		cols = ("user", "password", "secret", "source", "found")
		tv = ttk.Treeview(parent, columns=cols, show="headings", height=10)
		for c, w in zip(cols, (120, 140, 200, 220, 140)):
			tv.heading(c, text=c.title())
			tv.column(c, width=w, anchor="w")
		tv.pack(fill="both", expand=True, padx=4, pady=4)
		self.tracker_creds = tv

		row = tk.Frame(parent, bg=BG)
		row.pack(fill="x", padx=4, pady=2)
		ttk.Label(row, text="User:").pack(side="left")
		e_user = ttk.Entry(row, width=14)
		e_user.pack(side="left", padx=2)
		ttk.Label(row, text="Pass:").pack(side="left")
		e_pass = ttk.Entry(row, width=14)
		e_pass.pack(side="left", padx=2)
		ttk.Label(row, text="Secret/Hash:").pack(side="left")
		e_sec = ttk.Entry(row, width=18)
		e_sec.pack(side="left", padx=2)
		ttk.Label(row, text="Source:").pack(side="left")
		e_src = ttk.Entry(row, width=18)
		e_src.pack(side="left", padx=2)
		def add():
			if not e_user.get() and not e_pass.get() and not e_sec.get():
				return
			if self.tracker_machine_id is None:
				return
			self.store.add_cred(self.tracker_machine_id, e_user.get(), e_pass.get(), e_sec.get(), e_src.get())
			e_user.delete(0, tk.END); e_pass.delete(0, tk.END); e_sec.delete(0, tk.END); e_src.delete(0, tk.END)
			self._tracker_refresh_creds()
		ttk.Button(row, text="Add", command=add).pack(side="left", padx=4)
		def remove():
			sel = tv.selection()
			if sel and messagebox.askyesno("Delete cred?", f"Delete {len(sel)} credential(s)?", parent=self.root):
				for s in sel:
					self.store.delete_cred(int(s))
				self._tracker_refresh_creds()
		ttk.Button(row, text="Delete", command=remove).pack(side="left", padx=2)

	def _build_tracker_loot(self, parent):
		cols = ("name", "kind", "value", "captured")
		tv = ttk.Treeview(parent, columns=cols, show="headings", height=10)
		for c, w in zip(cols, (140, 100, 320, 140)):
			tv.heading(c, text=c.title())
			tv.column(c, width=w, anchor="w")
		tv.pack(fill="both", expand=True, padx=4, pady=4)
		self.tracker_loot = tv

		row = tk.Frame(parent, bg=BG)
		row.pack(fill="x", padx=4, pady=2)
		ttk.Label(row, text="Name:").pack(side="left")
		e_name = ttk.Entry(row, width=18); e_name.pack(side="left", padx=2)
		ttk.Label(row, text="Kind:").pack(side="left")
		e_kind = ttk.Combobox(row, values=["hash", "file", "screenshot", "creds", "key", "binary", "ticket", "other"], width=10)
		e_kind.pack(side="left", padx=2)
		ttk.Label(row, text="Value:").pack(side="left")
		e_val = ttk.Entry(row, width=40); e_val.pack(side="left", padx=2)
		def add():
			if not e_name.get() or not e_val.get() or self.tracker_machine_id is None:
				return
			self.store.add_loot(self.tracker_machine_id, e_name.get(), e_kind.get() or "other", e_val.get())
			e_name.delete(0, tk.END); e_val.delete(0, tk.END)
			self._tracker_refresh_loot()
		ttk.Button(row, text="Add", command=add).pack(side="left", padx=4)
		def remove():
			sel = tv.selection()
			if sel and messagebox.askyesno("Delete loot?", f"Delete {len(sel)} loot entries?", parent=self.root):
				for s in sel:
					self.store.delete_loot(int(s))
				self._tracker_refresh_loot()
		ttk.Button(row, text="Delete", command=remove).pack(side="left", padx=2)

	def _build_tracker_screens(self, parent):
		cols = ("label", "path", "taken")
		tv = ttk.Treeview(parent, columns=cols, show="headings", height=10)
		for c, w in zip(cols, (140, 360, 140)):
			tv.heading(c, text=c.title())
			tv.column(c, width=w, anchor="w")
		tv.pack(fill="both", expand=True, padx=4, pady=4)
		self.tracker_screens = tv
		tv.bind("<Double-1>", self._tracker_open_screenshot)

		row = tk.Frame(parent, bg=BG)
		row.pack(fill="x", padx=4, pady=2)
		ttk.Label(row, text="Label:").pack(side="left")
		e_lbl = ttk.Entry(row, width=20); e_lbl.pack(side="left", padx=2)
		ttk.Button(row, text="Browse file...", command=lambda: self._tracker_browse_screenshot(e_lbl)).pack(side="left", padx=2)
		def remove():
			sel = tv.selection()
			if sel and messagebox.askyesno("Delete?", f"Delete {len(sel)} screenshot link(s)?", parent=self.root):
				for s in sel:
					self.store.delete_screen(int(s))
				self._tracker_refresh_screens()
		ttk.Button(row, text="Delete", command=remove).pack(side="left", padx=2)

	def _build_tracker_checklist(self, parent):
		# Heading + scrollable text with checkboxes per phase
		canvas_frame = tk.Frame(parent, bg=BG)
		canvas_frame.pack(fill="both", expand=True, padx=4, pady=4)
		canvas = tk.Canvas(canvas_frame, bg=BG, highlightthickness=0)
		scroll = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
		canvas.configure(yscrollcommand=scroll.set)
		scroll.pack(side="right", fill="y")
		canvas.pack(side="left", fill="both", expand=True)
		inner = tk.Frame(canvas, bg=BG)
		canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
		inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
		canvas.bind("<Configure>", lambda e: canvas.itemconfig("inner", width=e.width))
		self.tracker_checklist_inner = inner

		footer = tk.Frame(parent, bg=BG)
		footer.pack(fill="x", padx=4, pady=4)
		ttk.Label(footer, text="Progress:").pack(side="left")
		self.tracker_progress_var = tk.StringVar(value="0/0")
		ttk.Label(footer, textvariable=self.tracker_progress_var, font=(MONO_FONT, 11, "bold"), foreground=ACCENT).pack(side="left", padx=4)

	def _build_tracker_flags(self, parent):
		w = tk.Text(parent, bg=BG, fg=FG, insertbackground=FG, font=(MONO_FONT, 12), wrap="word", height=8)
		w.pack(fill="both", expand=True, padx=4, pady=4)
		w.bind("<FocusOut>", lambda e: self._tracker_save_flags())
		w.bind("<KeyRelease>", lambda e: self._tracker_save_flags())
		self.tracker_flags = w

	def _build_tracker_notes(self, parent):
		w = tk.Text(parent, bg=BG, fg=FG, insertbackground=FG, font=(UI_FONT, 11), wrap="word", height=10)
		w.pack(fill="both", expand=True, padx=4, pady=4)
		w.bind("<FocusOut>", lambda e: self._tracker_save_notes())
		self.tracker_machine_notes = w

	def _tracker_refresh(self):
		if not hasattr(self, "machine_list"):
			return
		self.machine_list.delete(0, tk.END)
		for m in self.store.list_machines():
			label = f"{m['name']:<24} {m['status']:<8} {m['difficulty']:<8} {m['address']}"
			self.machine_list.insert(tk.END, label)
		# Map listbox index -> machine id
		self._machine_ids = [m["id"] for m in self.store.list_machines()]
		if self._machine_ids and self.tracker_machine_id not in self._machine_ids:
			self.machine_list.selection_clear(0, tk.END)
			self.machine_list.selection_set(0)
			self.tracker_machine_id = self._machine_ids[0]
		self._tracker_load_details()

	def _tracker_on_select(self, event):
		sel = self.machine_list.curselection()
		if not sel:
			return
		idx = sel[0]
		if idx < len(self._machine_ids):
			self.tracker_machine_id = self._machine_ids[idx]
			self._tracker_load_details()

	def _tracker_load_details(self):
		if self.tracker_machine_id is None:
			return
		m = self.store.get_machine(self.tracker_machine_id)
		if not m:
			return
		self.m_name_var.set(m["name"] or "")
		self.m_addr_var.set(m["address"] or "")
		self.m_diff_var.set(m["difficulty"] or "unknown")
		self.m_status_var.set(m["status"] or "active")
		self._tracker_refresh_creds()
		self._tracker_refresh_loot()
		self._tracker_refresh_screens()
		self._tracker_refresh_checklist()
		self._tracker_refresh_flags()
		self._tracker_refresh_notes()
		self._tracker_update_elapsed()

	def _tracker_save_machine(self):
		if self.tracker_machine_id is None:
			return
		self.store.update_machine(
			self.tracker_machine_id,
			name=self.m_name_var.get(),
			address=self.m_addr_var.get(),
			difficulty=self.m_diff_var.get(),
			status=self.m_status_var.get(),
		)
		self._flash_status("Machine saved")
		self._tracker_refresh()

	def _tracker_new_machine(self):
		name = simpledialog.askstring("New machine", "Name:", parent=self.root)
		if not name:
			return
		methodology = None
		try:
			import cheatseeds
			methodology = cheatseeds.METHODOLOGY
		except Exception:
			methodology = None
		mid = self.store.create_machine(name, methodology=methodology)
		self.tracker_machine_id = mid
		self._tracker_refresh()
		self._flash_status(f"Created '{name}'")

	def _tracker_toggle_pause(self):
		if self.tracker_machine_id is None:
			return
		m = self.store.get_machine(self.tracker_machine_id)
		if not m:
			return
		if m["status"] == "paused":
			# Resume
			import time as _t
			now_iso = datetime.now().isoformat(timespec="seconds")
			self.store.update_machine(self.tracker_machine_id, started_at=now_iso, status="active")
		else:
			# Pause - accumulate
			import time as _t
			try:
				started = datetime.fromisoformat(m["started_at"])
				delta = (datetime.now() - started).total_seconds()
			except Exception:
				delta = 0
			paused = (m["paused_secs"] or 0) + int(delta)
			self.store.update_machine(self.tracker_machine_id, paused_secs=paused, started_at=None, status="paused")
		self._tracker_load_details()
		self._flash_status(f"Status: {self.m_status_var.get()}")

	def _tracker_set_status(self, st):
		if self.tracker_machine_id is None:
			return
		self.store.update_machine(self.tracker_machine_id, status=st)
		self.m_status_var.set(st)
		self._tracker_refresh()

	def _tracker_delete_machine(self):
		if self.tracker_machine_id is None:
			return
		if not messagebox.askyesno("Delete machine?", "This deletes the machine and ALL its creds/loot/screenshots/checklist.", parent=self.root):
			return
		self.store.delete_machine(self.tracker_machine_id)
		self.tracker_machine_id = None
		self._tracker_refresh()
		self._flash_status("Machine deleted")

	def _tracker_refresh_creds(self):
		if not hasattr(self, "tracker_creds"):
			return
		tv = self.tracker_creds
		tv.delete(*tv.get_children())
		for r in self.store.list_creds(self.tracker_machine_id):
			tv.insert("", tk.END, iid=str(r["id"]), values=(r["username"], r["password"], r["secret"], r["source"], r["found_at"]))

	def _tracker_refresh_loot(self):
		if not hasattr(self, "tracker_loot"):
			return
		tv = self.tracker_loot
		tv.delete(*tv.get_children())
		for r in self.store.list_loot(self.tracker_machine_id):
			tv.insert("", tk.END, iid=str(r["id"]), values=(r["name"], r["kind"], r["value"], r["captured_at"]))

	def _tracker_refresh_screens(self):
		if not hasattr(self, "tracker_screens"):
			return
		tv = self.tracker_screens
		tv.delete(*tv.get_children())
		for r in self.store.list_screens(self.tracker_machine_id):
			tv.insert("", tk.END, iid=str(r["id"]), values=(r["label"], r["path"], r["taken_at"]))

	def _tracker_refresh_checklist(self):
		if not hasattr(self, "tracker_checklist_inner"):
			return
		# Wipe
		for w in self.tracker_checklist_inner.winfo_children():
			w.destroy()
		rows = self.store.list_checklist(self.tracker_machine_id)
		# group by phase
		by_phase: dict = {}
		for r in rows:
			by_phase.setdefault(r["phase"], []).append(r)
		done = total = 0
		for phase in sorted(by_phase.keys()):
			ttk.Label(self.tracker_checklist_inner, text=phase.title(), background=BG, foreground=ACCENT, font=(UI_FONT, 12, "bold")).pack(anchor="w", padx=4, pady=(8, 2))
			for r in by_phase[phase]:
				var = tk.IntVar(value=r["ticked"])
				def tick(v=var, rid=r["id"]):
					self.store.tick_checklist(rid, bool(v.get()))
					self._tracker_update_progress()
				chk = ttk.Checkbutton(self.tracker_checklist_inner, text=r["item"], variable=var, command=tick)
				chk.pack(anchor="w", padx=20, pady=1)
				total += 1
				if r["ticked"]:
					done += 1
		self.tracker_done = done
		self.tracker_total = total
		self._tracker_update_progress()

	def _tracker_update_progress(self):
		if not hasattr(self, "tracker_progress_var"):
			return
		done, total = self.store.checklist_progress(self.tracker_machine_id)
		self.tracker_progress_var.set(f"{done}/{total}")

	def _tracker_refresh_flags(self):
		if not hasattr(self, "tracker_flags"):
			return
		m = self.store.get_machine(self.tracker_machine_id)
		if not m:
			return
		self._suppress_change = True
		self.tracker_flags.delete("1.0", tk.END)
		self.tracker_flags.insert("1.0", m["flags"] or "")
		self._suppress_change = False

	def _tracker_save_flags(self):
		if self._suppress_change or self.tracker_machine_id is None:
			return
		text = self.tracker_flags.get("1.0", tk.END)
		self.store.update_machine(self.tracker_machine_id, flags=text)

	def _tracker_refresh_notes(self):
		if not hasattr(self, "tracker_machine_notes"):
			return
		m = self.store.get_machine(self.tracker_machine_id)
		if not m:
			return
		self._suppress_change = True
		self.tracker_machine_notes.delete("1.0", tk.END)
		self.tracker_machine_notes.insert("1.0", m["notes"] or "")
		self._suppress_change = False

	def _tracker_save_notes(self):
		if self._suppress_change or self.tracker_machine_id is None:
			return
		text = self.tracker_machine_notes.get("1.0", tk.END)
		self.store.update_machine(self.tracker_machine_id, notes=text)

	def _tracker_browse_screenshot(self, label_entry):
		if self.tracker_machine_id is None:
			return
		path = filedialog.askopenfilename(title="Select screenshot", parent=self.root,
			filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.gif *.webp"), ("All", "*.*")])
		if not path:
			return
		label = label_entry.get() or Path(path).stem
		self.store.add_screen(self.tracker_machine_id, label, path)
		self._tracker_refresh_screens()
		self._flash_status(f"Added screenshot: {label}")

	def _tracker_open_screenshot(self, event):
		sel = self.tracker_screens.selection()
		if not sel:
			return
		path = self.store.list_screens(self.tracker_machine_id)[int(sel[0]) - 1]["path"] if False else None
		# safer: re-query
		rows = self.store.list_screens(self.tracker_machine_id)
		for r in rows:
			if str(r["id"]) == sel[0]:
				path = r["path"]
				break
		if not path:
			return
		try:
			webbrowser.open(Path(path).as_uri())
		except Exception as e:
			messagebox.showerror("Open failed", str(e), parent=self.root)

	def _tracker_update_elapsed(self):
		if not hasattr(self, "m_elapsed_var") or self.tracker_machine_id is None:
			return
		m = self.store.get_machine(self.tracker_machine_id)
		if not m["status"] == "owned" or m["status"] == "retired":
			self.m_elapsed_var.set("done")
			return
		try:
			if m["status"] == "paused":
				secs = int(m["paused_secs"] or 0)
			else:
				started = datetime.fromisoformat(m["started_at"])
				delta = (datetime.now() - started).total_seconds()
				secs = int((m["paused_secs"] or 0) + delta)
			h, r = divmod(secs, 3600); m_, s = divmod(r, 60)
			self.m_elapsed_var.set(f"{h:02d}:{m_:02d}:{s:02d}")
		except Exception:
			self.m_elapsed_var.set("--:--:--")

	def _schedule_timer_tick(self):
		try:
			self._tracker_update_elapsed()
		except Exception:
			pass
		try:
			self.root.after(1000, self._schedule_timer_tick)
		except Exception:
			pass

	# ----------------------------- payloads tab ----------------------------- #

	def _build_checklist_tab(self):
		"""LainKusanagi OSCP practice list — 255 boxes across HTB, HackSmarter, PG, VulnLab, VulnHub.

		Treeview with columns: Done, Name, Platform, OS, Kind, Status, Tracker.
		Toolbar: platform filter, status filter, search, stats, refresh, mark-all.
		"""
		frame = tk.Frame(self.notebook, bg=BG)
		self.notebook.add(frame, text="Checklist")
		self._tab_frames["checklist"] = frame
		self._checklist_built = True

		# Toolbar
		toolbar = tk.Frame(frame, bg=BG)
		toolbar.pack(fill="x", padx=8, pady=(8, 4))

		tk.Label(toolbar, text="Platform:", bg=BG, fg=FG, font=(UI_FONT, 10)).pack(side="left", padx=(0, 4))
		self.checklist_platform_var = tk.StringVar(value="All")
		platforms = ["All"] + self.store.checklist_platforms()
		self.checklist_platform_cb = ttk.Combobox(toolbar, textvariable=self.checklist_platform_var,
			values=platforms, state="readonly", width=20)
		self.checklist_platform_cb.pack(side="left", padx=(0, 12))
		self.checklist_platform_cb.bind("<<ComboboxSelected>>", lambda e: self._checklist_refresh())

		tk.Label(toolbar, text="Status:", bg=BG, fg=FG, font=(UI_FONT, 10)).pack(side="left", padx=(0, 4))
		self.checklist_status_var = tk.StringVar(value="All")
		statuses = ["all", "locked", "started", "user", "root", "owned", "skipped"]
		status_labels = {"all": "All", "locked": "Locked", "started": "Started", "user": "User",
			"root": "Root", "owned": "Owned", "skipped": "Skipped"}
		self.checklist_status_cb = ttk.Combobox(toolbar, textvariable=self.checklist_status_var,
			values=[status_labels[s] for s in statuses], state="readonly", width=12)
		self.checklist_status_cb.pack(side="left", padx=(0, 12))
		# Map display label -> status key
		self._checklist_status_map = {status_labels[s]: s for s in statuses}
		self.checklist_status_cb.bind("<<ComboboxSelected>>", lambda e: self._checklist_refresh())

		tk.Label(toolbar, text="Search:", bg=BG, fg=FG, font=(UI_FONT, 10)).pack(side="left", padx=(0, 4))
		self.checklist_search_var = tk.StringVar()
		search_entry = ttk.Entry(toolbar, textvariable=self.checklist_search_var, width=20)
		search_entry.pack(side="left", padx=(0, 12))
		self.checklist_search_var.trace_add("write", lambda *a: self._checklist_refresh())

		ttk.Button(toolbar, text="Show only unfinished",
			command=lambda: self._checklist_filter_unfinished()).pack(side="left", padx=4)
		ttk.Button(toolbar, text="Reset filters", command=self._checklist_reset_filters).pack(side="left", padx=4)
		ttk.Button(toolbar, text="Refresh", command=self._checklist_refresh).pack(side="left", padx=4)

		# Treeview
		tree_wrap = tk.Frame(frame, bg=BG)
		tree_wrap.pack(fill="both", expand=True, padx=8, pady=4)

		cols = ("done", "name", "platform", "os", "kind", "status", "tracker")
		tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", selectmode="browse")
		tree.heading("done", text="Done")
		tree.heading("name", text="Machine")
		tree.heading("platform", text="Platform")
		tree.heading("os", text="OS")
		tree.heading("kind", text="Category")
		tree.heading("status", text="Status")
		tree.heading("tracker", text="Tracker")
		tree.column("done", width=60, anchor="center")
		tree.column("name", width=200)
		tree.column("platform", width=160)
		tree.column("os", width=80, anchor="center")
		tree.column("kind", width=160)
		tree.column("status", width=100, anchor="center")
		tree.column("tracker", width=80, anchor="center")
		vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=tree.yview)
		hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=tree.xview)
		tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
		tree.grid(row=0, column=0, sticky="nsew")
		vsb.grid(row=0, column=1, sticky="ns")
		hsb.grid(row=1, column=0, sticky="ew")
		tree_wrap.grid_rowconfigure(0, weight=1)
		tree_wrap.grid_columnconfigure(0, weight=1)
		self.checklist_tree = tree
		# Color-code the status column when items are rendered
		tree.tag_configure("owned", foreground="#22c55e")
		tree.tag_configure("root", foreground="#22c55e")
		tree.tag_configure("user", foreground="#eab308")
		tree.tag_configure("started", foreground="#f59e0b")
		tree.tag_configure("skipped", foreground="#6b7280")
		tree.tag_configure("locked", foreground=FG)
		tree.bind("<Double-1>", self._checklist_on_dblclick)
		tree.bind("<Button-3>", self._checklist_on_rightclick) # right-click menu on Mac via Ctrl+click

		# Bottom bar
		bottom = tk.Frame(frame, bg=BG)
		bottom.pack(fill="x", padx=8, pady=4)
		self.checklist_stats_label = tk.Label(bottom, text="", bg=BG, fg=FG,
			font=(UI_FONT, 10), anchor="w")
		self.checklist_stats_label.pack(side="left", fill="x", expand=True)
		ttk.Button(bottom, text="Mark Selected: Started",
			command=lambda: self._checklist_set_status("started")).pack(side="left", padx=2)
		ttk.Button(bottom, text="User",
			command=lambda: self._checklist_set_status("user")).pack(side="left", padx=2)
		ttk.Button(bottom, text="Root",
			command=lambda: self._checklist_set_status("root")).pack(side="left", padx=2)
		ttk.Button(bottom, text="Owned",
			command=lambda: self._checklist_set_status("owned")).pack(side="left", padx=2)
		ttk.Button(bottom, text="Skip",
			command=lambda: self._checklist_set_status("skipped")).pack(side="left", padx=2)
		ttk.Button(bottom, text="Reset to Locked",
			command=lambda: self._checklist_set_status("locked")).pack(side="left", padx=2)

		self._checklist_refresh()

	def _checklist_reset_filters(self):
		self.checklist_platform_var.set("All")
		self.checklist_status_var.set("All")
		self.checklist_status_cb.set("All")
		self.checklist_search_var.set("")
		self._checklist_refresh()

	def _checklist_filter_unfinished(self):
		self.checklist_status_var.set("Locked")
		self.checklist_status_cb.set("Locked")
		self._checklist_refresh()

	def _checklist_refresh(self):
		if not getattr(self, "_checklist_built", False):
			return
		platform = self.checklist_platform_var.get()
		platform_key = platform if platform != "All" else "all"
		status_label = self.checklist_status_var.get()
		status_key = self._checklist_status_map.get(status_label, "all")
		search = self.checklist_search_var.get().strip() or None
		rows = self.store.list_checklist_items(platform=platform_key, status=status_key, search=search)
		# Preserve selection
		sel = self.checklist_tree.selection()
		prev_id = sel[0] if sel else None
		for iid in self.checklist_tree.get_children():
			self.checklist_tree.delete(iid)
		for r in rows:
			item_id, plat, os_name, kind, name, status, notes, machine_id = r
			done = "yes" if status in ("owned", "root") else ""
			tracker = "yes" if machine_id else ""
			tags = (status,)
			self.checklist_tree.insert("", "end", iid=str(item_id), values=(
				done, name, plat, os_name, kind, status, tracker), tags=tags)
		if prev_id and self.checklist_tree.exists(prev_id):
			self.checklist_tree.selection_set(prev_id)
		# Update stats
		s = self.store.checklist_stats()
		parts = [f"Total: {s['total']}"]
		for k in ("owned", "root", "user", "started", "locked", "skipped"):
			if s["by_status"].get(k):
				parts.append(f"{k}: {s['by_status'][k]}")
		self.checklist_stats_label.config(text=" | ".join(parts))

	def _checklist_selected_id(self):
		sel = self.checklist_tree.selection()
		if not sel:
			return None
		return int(sel[0])

	def _checklist_set_status(self, status: str):
		item_id = self._checklist_selected_id()
		if item_id is None:
			return
		self.store.update_checklist_status(item_id, status)
		self._checklist_refresh()
		self._flash_status(f"Status -> {status}")

	def _checklist_on_dblclick(self, event):
		# Double-click cycles status: locked -> started -> user -> root -> owned -> locked
		item_id = self._checklist_selected_id()
		if item_id is None:
			return
		row = next((r for r in self.store.list_checklist_items() if r[0] == item_id), None)
		if not row:
			return
		order = ["locked", "started", "user", "root", "owned"]
		cur = row[5]
		idx = order.index(cur) if cur in order else 0
		nxt = order[(idx + 1) % len(order)]
		self.store.update_checklist_status(item_id, nxt)
		self._checklist_refresh()

	def _checklist_on_rightclick(self, event):
		item_id = self._checklist_tree_row_at(event)
		if item_id is None:
			return
		self.checklist_tree.selection_set(item_id)
		menu = tk.Menu(self.root, tearoff=0)
		menu.add_command(label="Set: Started", command=lambda: self._checklist_set_status("started"))
		menu.add_command(label="Set: User", command=lambda: self._checklist_set_status("user"))
		menu.add_command(label="Set: Root", command=lambda: self._checklist_set_status("root"))
		menu.add_command(label="Set: Owned", command=lambda: self._checklist_set_status("owned"))
		menu.add_command(label="Set: Skipped", command=lambda: self._checklist_set_status("skipped"))
		menu.add_command(label="Set: Locked (reset)", command=lambda: self._checklist_set_status("locked"))
		menu.add_separator()
		menu.add_command(label="Open as Tracker Machine...", command=self._checklist_open_as_tracker)
		menu.add_command(label="Edit notes...", command=self._checklist_edit_notes)
		menu.add_separator()
		menu.add_command(label="Open ippsec search",
			command=self._checklist_open_ippsec)
		try:
			menu.tk_popup(event.x_root, event.y_root)
		except Exception:
			pass

	def _checklist_tree_row_at(self, event):
		iid = self.checklist_tree.identify_row(event.y)
		return int(iid) if iid else None

	def _checklist_open_ippsec(self):
		item_id = self._checklist_selected_id()
		if item_id is None:
			return
		row = next((r for r in self.store.list_checklist_items() if r[0] == item_id), None)
		if not row:
			return
		name = row[4]
		from urllib.parse import quote_plus
		url = f"https://www.google.com/search?q=ippsec+{quote_plus(name)}+walkthrough"
		import webbrowser as _wb
		_wb.open(url)
		self._flash_status(f"Opened ippsec search for {name}")

	def _checklist_edit_notes(self):
		item_id = self._checklist_selected_id()
		if item_id is None:
			return
		row = next((r for r in self.store.list_checklist_items() if r[0] == item_id), None)
		if not row:
			return
		import tkinter.simpledialog as sd
		current = row[6] or ""
		notes = sd.askstring("Checklist notes", f"Notes for {row[4]}:", initialvalue=current, parent=self.root)
		if notes is not None:
			self.store.update_checklist_notes(item_id, notes)
			self._flash_status("Notes saved")

	def _checklist_open_as_tracker(self):
		"""Create a Tracker machine from the selected checklist entry, link them, switch to Tracker tab."""
		item_id = self._checklist_selected_id()
		if item_id is None:
			return
		row = next((r for r in self.store.list_checklist_items() if r[0] == item_id), None)
		if not row:
			return
		*_, name, status, notes, machine_id = row
		if machine_id:
			# Already linked — just switch to it
			self._select_tab("tracker")
			self.tracker_machine_id = machine_id
			self._tracker_refresh()
			try:
				idx = self._machine_ids.index(machine_id)
				self.machine_list.selection_clear(0, tk.END)
				self.machine_list.selection_set(idx)
				self.machine_list.see(idx)
				self._tracker_on_select()
			except Exception:
				pass
			return
		try:
			import cheatseeds
			methodology = cheatseeds.METHODOLOGY
		except Exception:
			methodology = None
		mid = self.store.create_machine(name=name, difficulty=row[3],
			methodology=methodology,
			notes=f"From LainKusanagi checklist ({row[1]} / {row[2]})\n{notes or ''}")
		self.store.update_checklist_machine(item_id, mid)
		self._flash_status(f"Created Tracker machine: {name}")
		self._select_tab("tracker")
		self.tracker_machine_id = mid
		self._tracker_refresh()
		try:
			idx = self._machine_ids.index(mid)
			self.machine_list.selection_clear(0, tk.END)
			self.machine_list.selection_set(idx)
			self.machine_list.see(idx)
			self._tracker_on_select()
		except Exception:
			pass

	def _build_payloads_tab(self):
		frame = tk.Frame(self.notebook, bg=BG)
		self.notebook.add(frame, text="Payloads")
		self._tab_frames["payloads"] = frame
		self._payloads_built = True

		# Top: search + filter
		top = tk.Frame(frame, bg=BG)
		top.pack(fill="x", padx=8, pady=8)
		ttk.Label(top, text="Search:", background=BG, foreground=FG_DIM).pack(side="left")
		self.payload_search_var = tk.StringVar()
		self.payload_search_var.trace_add("write", lambda *_: self._payloads_refresh())
		ttk.Entry(top, textvariable=self.payload_search_var, width=24, font=(UI_FONT, 12)).pack(side="left", padx=4)
		ttk.Label(top, text="Category:", background=BG, foreground=FG_DIM).pack(side="left", padx=(12, 4))
		self.payload_cat_var = tk.StringVar(value="all")
		cats = ["all"] + self.store.list_payload_categories()
		cmb = ttk.Combobox(top, textvariable=self.payload_cat_var, values=cats, state="readonly", width=18)
		cmb.pack(side="left")
		cmb.bind("<<ComboboxSelected>>", lambda e: self._payloads_refresh())
		ttk.Button(top, text="+ New Payload", command=self._payload_new).pack(side="left", padx=8)
		ttk.Button(top, text="Refresh", command=self._payloads_refresh).pack(side="left", padx=2)

		# Split: list / detail
		body = tk.PanedWindow(frame, orient="horizontal", sashwidth=6, background=BG, borderwidth=0)
		body.pack(fill="both", expand=True, padx=8, pady=(0, 8))

		left = tk.Frame(body, bg=BG_ALT)
		body.add(left, minsize=240, width=320)
		cols = ("name", "category", "platform")
		tv = ttk.Treeview(left, columns=cols, show="headings")
		for c, w in zip(cols, (180, 110, 90)):
			tv.heading(c, text=c.title())
			tv.column(c, width=w, anchor="w")
		tv.pack(fill="both", expand=True, padx=4, pady=4)
		tv.bind("<<TreeviewSelect>>", self._payloads_on_select)
		self.payloads_list = tv
		self._payload_ids = []

		right = tk.Frame(body, bg=BG)
		body.add(right, minsize=320)
		self.payloads_detail = tk.Text(right, bg=BG, fg=FG, insertbackground=FG,
			font=(MONO_FONT, 12), wrap="word", padx=10, pady=10, relief="flat", height=14)
		self.payloads_detail.pack(fill="both", expand=True, padx=4, pady=4)
		btn_row = tk.Frame(right, bg=BG)
		btn_row.pack(fill="x", padx=4, pady=4)
		ttk.Button(btn_row, text="Copy", command=self._payload_copy).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Insert into current note", command=self._payload_insert).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Edit", command=self._payload_edit).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Delete", command=self._payload_delete).pack(side="left", padx=2)

		self._payloads_refresh()

	def _payloads_refresh(self):
		if not hasattr(self, "payloads_list"):
			return
		q = self.payload_search_var.get() if hasattr(self, "payload_search_var") else ""
		cat = self.payload_cat_var.get() if hasattr(self, "payload_cat_var") else "all"
		if q:
			rows = self.store.search_payloads(q)
		else:
			rows = self.store.list_payloads(category=cat)
		tv = self.payloads_list
		tv.delete(*tv.get_children())
		self._payload_ids = []
		for r in rows:
			tv.insert("", tk.END, iid=str(r["id"]), values=(r["name"], r["category"], r["platform"]))
			self._payload_ids.append(r["id"])

	def _payloads_on_select(self, event):
		sel = self.payloads_list.selection()
		if not sel:
			return
		pid = int(sel[0])
		p = self.store.get_payload(pid)
		if not p:
			return
		self.payloads_detail.delete("1.0", tk.END)
		header = f"# {p['name']} ({p['category']} / {p['platform']})\n# tags: {p['tags']}\n\n"
		self.payloads_detail.insert("1.0", header + p["content"])

	def _payload_copy(self):
		text = self.payloads_detail.get("1.0", tk.END)
		# Strip header
		lines = text.splitlines()
		body = "\n".join(line for line in lines if not line.startswith("# "))
		self.root.clipboard_clear()
		self.root.clipboard_append(body)
		self._flash_status("Payload copied to clipboard")

	def _payload_insert(self):
		sel = self.payloads_list.selection()
		if not sel:
			return
		p = self.store.get_payload(int(sel[0]))
		if not p:
			return
		# Switch to Notes tab
		self._select_tab("notes")
		# Need a current note - if none, create one
		if self.current_id is None:
			nid = self.store.create(title=f"Payloads - {p['name']}", phase="exploit", tags="payload")
			self._refresh_tree()
			self._select_note(nid)
		else:
			# Append a fenced code block to current note
			old = self.editor.get("1.0", tk.END)
			insertion = f"\n\n## {p['name']}\n\n```\n{p['content']}\n```\n"
			self.editor.insert(tk.END, insertion)
			self._on_change()
		self._flash_status(f"Inserted: {p['name']}")

	def _payload_edit(self):
		sel = self.payloads_list.selection()
		if not sel:
			return
		pid = int(sel[0])
		p = self.store.get_payload(pid)
		if not p:
			return
		new_content = self.payloads_detail.get("1.0", tk.END)
		# Strip header (lines starting with "# " before blank)
		if "\n\n" in new_content:
			new_content = new_content.split("\n\n", 1)[1]
		self.store.update_payload(pid, content=new_content.rstrip())
		self._flash_status("Payload updated")

	def _payload_new(self):
		d = _PayloadDialog(self.root, self.store)
		self.root.wait_window(d.window)
		if d.result_id:
			self._payloads_refresh()
			# Select the new one
			for iid in self.payloads_list.get_children():
				if int(iid) == d.result_id:
					self.payloads_list.selection_set(iid)
					self.payloads_list.see(iid)
					break

	def _payload_delete(self):
		sel = self.payloads_list.selection()
		if not sel:
			return
		if not messagebox.askyesno("Delete payload?", "Delete this payload?", parent=self.root):
			return
		self.store.delete_payload(int(sel[0]))
		self._payloads_refresh()
		self._flash_status("Payload deleted")

	# ----------------------------- vault tab ----------------------------- #

	def _build_vault_tab(self):
		frame = tk.Frame(self.notebook, bg=BG)
		self.notebook.add(frame, text="Vault")
		self._tab_frames["vault"] = frame
		self._vault_built = True

		# Vault state
		self._vault_key = None # Fernet key in memory
		self._vault_salt = None
		self._vault_reveal_job = None

		# Check if vault is set up (salt exists in prefs)
		salt_b64 = self.store.get_pref("vault_salt", "")
		if salt_b64:
			# Show unlock UI
			self._vault_show_unlock()
		else:
			self._vault_show_setup()

	def _vault_show_setup(self):
		# Clear vault frame
		frame = self._tab_frames["vault"]
		for w in frame.winfo_children():
			w.destroy()
		# Form: set master password
		ttk.Label(frame, text="Set up your encrypted vault", background=BG, foreground=ACCENT,
			font=(UI_FONT, 14, "bold")).pack(pady=(30, 8))
		ttk.Label(frame, text="Pick a master password. If you forget it, your data is unrecoverable.",
			background=BG, foreground=FG_DIM, wraplength=520).pack(pady=(0, 12))
		row1 = tk.Frame(frame, bg=BG); row1.pack(pady=4)
		ttk.Label(row1, text="Master password:", background=BG, foreground=FG_DIM).pack(side="left", padx=4)
		e1 = ttk.Entry(row1, show="*", width=24, font=(UI_FONT, 12)); e1.pack(side="left", padx=4)
		row2 = tk.Frame(frame, bg=BG); row2.pack(pady=4)
		ttk.Label(row2, text="Confirm: ", background=BG, foreground=FG_DIM).pack(side="left", padx=4)
		e2 = ttk.Entry(row2, show="*", width=24, font=(UI_FONT, 12)); e2.pack(side="left", padx=4)
		def do_setup():
			p1, p2 = e1.get(), e2.get()
			if not p1:
				messagebox.showwarning("Empty password", "Password cannot be empty.", parent=self.root)
				return
			if p1 != p2:
				messagebox.showerror("Mismatch", "Passwords do not match.", parent=self.root)
				return
			if len(p1) < 6:
				if not messagebox.askyesno("Weak password", "Use this short password anyway?", parent=self.root):
					return
			salt = os.urandom(16)
			self.store.set_pref("vault_salt", base64.b64encode(salt).decode("ascii"))
			self._vault_salt = salt
			self._vault_key = derive_vault_key(p1, salt)
			from cryptography.fernet import Fernet
			self._vault_fernet = Fernet(self._vault_key)
			self._flash_status("Vault created")
			self._vault_show_unlocked()
		ttk.Button(frame, text="Create vault", command=do_setup).pack(pady=12)

	def _vault_show_unlock(self):
		frame = self._tab_frames["vault"]
		for w in frame.winfo_children():
			w.destroy()
		ttk.Label(frame, text="Encrypted vault is locked", background=BG, foreground=ACCENT,
			font=(UI_FONT, 14, "bold")).pack(pady=(30, 8))
		ttk.Label(frame, text="Enter master password to unlock.", background=BG, foreground=FG_DIM).pack(pady=(0, 12))
		e = ttk.Entry(frame, show="*", width=24, font=(UI_FONT, 12))
		e.pack(pady=4)
		def do_unlock(event=None):
			pw = e.get()
			if not pw:
				return
			salt_b64 = self.store.get_pref("vault_salt", "")
			if not salt_b64:
				self._vault_show_setup()
				return
			self._vault_salt = base64.b64decode(salt_b64)
			key = derive_vault_key(pw, self._vault_salt)
			from cryptography.fernet import Fernet, InvalidToken
			fer = Fernet(key)
			# Quick verification: try decrypting the first row if any
			rows = self.store.list_vault()
			if rows:
				row = self.store.get_vault(rows[0]["id"])
				try:
					fer.decrypt(row["ciphertext"])
				except InvalidToken:
					messagebox.showerror("Wrong password", "Could not decrypt the vault. Wrong password?", parent=self.root)
					return
			self._vault_key = key
			self._vault_fernet = fer
			self._flash_status("Vault unlocked")
			self._vault_show_unlocked()
		ttk.Button(frame, text="Unlock", command=do_unlock).pack(pady=8)
		e.bind("<Return>", do_unlock)
		e.focus_set()

	def _vault_show_unlocked(self):
		frame = self._tab_frames["vault"]
		for w in frame.winfo_children():
			w.destroy()

		# Top bar
		top = tk.Frame(frame, bg=BG)
		top.pack(fill="x", padx=8, pady=8)
		ttk.Button(top, text="+ New entry", command=self._vault_new).pack(side="left", padx=2)
		ttk.Button(top, text="Lock vault", command=self._vault_lock).pack(side="left", padx=8)
		ttk.Label(top, text="Reveal sensitive text for:", background=BG, foreground=FG_DIM).pack(side="left", padx=(16, 4))
		self._vault_reveal_var = tk.IntVar(value=30)
		ttk.Spinbox(top, from_=0, to=600, increment=5, width=5, textvariable=self._vault_reveal_var).pack(side="left")
		ttk.Label(top, text="seconds (0 = off)", background=BG, foreground=FG_DIM).pack(side="left", padx=4)

		# List + detail
		body = tk.PanedWindow(frame, orient="horizontal", sashwidth=6, background=BG, borderwidth=0)
		body.pack(fill="both", expand=True, padx=8, pady=(0, 8))

		left = tk.Frame(body, bg=BG_ALT)
		body.add(left, minsize=220, width=280)
		cols = ("title", "updated", "size")
		tv = ttk.Treeview(left, columns=cols, show="headings")
		for c, w in zip(cols, (180, 130, 60)):
			tv.heading(c, text=c.title())
			tv.column(c, width=w, anchor="w")
		tv.pack(fill="both", expand=True, padx=4, pady=4)
		tv.bind("<<TreeviewSelect>>", self._vault_on_select)
		self.vault_list = tv

		right = tk.Frame(body, bg=BG)
		body.add(right, minsize=320)
		# Title entry
		ttk.Label(right, text="Title:", background=BG, foreground=FG_DIM).pack(anchor="w", padx=8, pady=(8, 0))
		self.vault_title_var = tk.StringVar()
		ttk.Entry(right, textvariable=self.vault_title_var, font=(UI_FONT, 12, "bold")).pack(fill="x", padx=8, pady=(0, 8))
		# Body text
		self.vault_body = tk.Text(right, bg=BG, fg=FG, insertbackground=FG,
			font=(MONO_FONT, 12), wrap="word", padx=10, pady=10, relief="flat")
		self.vault_body.pack(fill="both", expand=True, padx=8, pady=4)
		# Save / Delete / Reveal
		btn_row = tk.Frame(right, bg=BG)
		btn_row.pack(fill="x", padx=8, pady=8)
		ttk.Button(btn_row, text="Save", command=self._vault_save).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Delete", command=self._vault_delete).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Reveal for N sec", command=self._vault_reveal_temp).pack(side="left", padx=2)
		ttk.Button(btn_row, text="Copy body", command=self._vault_copy_body).pack(side="left", padx=2)

		self._vault_refresh()

	def _vault_refresh(self):
		if not hasattr(self, "vault_list"):
			return
		tv = self.vault_list
		tv.delete(*tv.get_children())
		for r in self.store.list_vault():
			tv.insert("", tk.END, iid=str(r["id"]), values=(r["title"], r["updated_at"], f"{r['size']}b"))

	def _vault_on_select(self, event):
		sel = self.vault_list.selection()
		if not sel:
			return
		row = self.store.get_vault(int(sel[0]))
		if not row:
			return
		self.vault_title_var.set(row["title"])
		try:
			plaintext = self._vault_fernet.decrypt(row["ciphertext"]).decode("utf-8")
		except Exception as e:
			plaintext = f"<decryption error: {e}>"
		self.vault_body.delete("1.0", tk.END)
		# Show masked by default
		masked = "".join("•" if ch not in "\n\t" else ch for ch in plaintext)
		self.vault_body.insert("1.0", masked)
		self._vault_current_id = int(sel[0])
		self._vault_plaintext = plaintext # keep in memory, do not save to DB

	def _vault_new(self):
		# Save current if dirty
		plaintext = "\n".join(self.vault_title_var.get().splitlines()) + "\n\n<new entry>"
		cipher = self._vault_fernet.encrypt(plaintext.encode("utf-8"))
		vid = self.store.create_vault(self.vault_title_var.get() or "Untitled", cipher)
		self._vault_refresh()
		self._flash_status("New entry created")

	def _vault_save(self):
		if not hasattr(self, "_vault_current_id") or self._vault_current_id is None:
			return
		plaintext = self.vault_body.get("1.0", tk.END).rstrip()
		cipher = self._vault_fernet.encrypt(plaintext.encode("utf-8"))
		self.store.update_vault(self._vault_current_id, self.vault_title_var.get(), cipher)
		# Re-mask
		masked = "".join("•" if ch not in "\n\t" else ch for ch in plaintext)
		self.vault_body.delete("1.0", tk.END)
		self.vault_body.insert("1.0", masked)
		self._vault_refresh()
		self._flash_status("Vault entry saved & re-encrypted")

	def _vault_delete(self):
		if not hasattr(self, "_vault_current_id") or self._vault_current_id is None:
			return
		if not messagebox.askyesno("Delete?", "Delete this vault entry?", parent=self.root):
			return
		self.store.delete_vault(self._vault_current_id)
		self._vault_current_id = None
		self.vault_title_var.set("")
		self.vault_body.delete("1.0", tk.END)
		self._vault_refresh()
		self._flash_status("Vault entry deleted")

	def _vault_reveal_temp(self):
		if not hasattr(self, "_vault_plaintext"):
			return
		# Replace masked text with real plaintext
		self.vault_body.delete("1.0", tk.END)
		self.vault_body.insert("1.0", self._vault_plaintext)
		secs = int(self._vault_reveal_var.get())
		if secs <= 0:
			return
		if self._vault_reveal_job is not None:
			try:
				self.root.after_cancel(self._vault_reveal_job)
			except Exception:
				pass
		def hide():
			plaintext = self._vault_plaintext
			masked = "".join("•" if ch not in "\n\t" else ch for ch in plaintext)
			self.vault_body.delete("1.0", tk.END)
			self.vault_body.insert("1.0", masked)
			self._flash_status("Re-hidden")
		self._vault_reveal_job = self.root.after(secs * 1000, hide)
		self._flash_status(f"Revealed for {secs}s")

	def _vault_copy_body(self):
		if not hasattr(self, "_vault_plaintext"):
			return
		self.root.clipboard_clear()
		self.root.clipboard_append(self._vault_plaintext)
		self._flash_status("Copied to clipboard")

	def _vault_lock(self):
		self._vault_key = None
		self._vault_fernet = None
		self._vault_plaintext = ""
		self._vault_current_id = None
		self._vault_show_unlock()
		self._flash_status("Vault locked")


# ----------------------------- helper dialogs ----------------------------- #

class _PayloadDialog:
	def __init__(self, parent, store):
		self.store = store
		self.result_id = None
		self.window = tk.Toplevel(parent)
		self.window.title("New payload")
		self.window.transient(parent)
		self.window.grab_set()
		self.window.configure(bg=BG)
		ttk.Label(self.window, text="Name:", background=BG, foreground=FG_DIM).grid(row=0, column=0, sticky="e", padx=8, pady=4)
		e_name = ttk.Entry(self.window, width=30)
		e_name.grid(row=0, column=1, padx=8, pady=4)
		ttk.Label(self.window, text="Category:", background=BG, foreground=FG_DIM).grid(row=1, column=0, sticky="e", padx=8, pady=4)
		e_cat = ttk.Combobox(self.window, values=["reverse-shell", "msfvenom", "webshell", "privesc", "pivot", "transfer", "ad", "enum", "evade", "other"], width=28)
		e_cat.grid(row=1, column=1, padx=8, pady=4)
		ttk.Label(self.window, text="Platform:", background=BG, foreground=FG_DIM).grid(row=2, column=0, sticky="e", padx=8, pady=4)
		e_plat = ttk.Combobox(self.window, values=["linux", "windows", "web", "ad", "any"], width=28)
		e_plat.grid(row=2, column=1, padx=8, pady=4)
		ttk.Label(self.window, text="Tags:", background=BG, foreground=FG_DIM).grid(row=3, column=0, sticky="e", padx=8, pady=4)
		e_tags = ttk.Entry(self.window, width=30)
		e_tags.grid(row=3, column=1, padx=8, pady=4)
		ttk.Label(self.window, text="Content:", background=BG, foreground=FG_DIM).grid(row=4, column=0, sticky="ne", padx=8, pady=4)
		t_content = tk.Text(self.window, width=60, height=10, font=(MONO_FONT, 11), bg=BG, fg=FG, insertbackground=FG)
		t_content.grid(row=4, column=1, padx=8, pady=4)
		def save():
			if not e_name.get():
				messagebox.showwarning("Missing name", "Name required.", parent=self.window)
				return
			if not t_content.get("1.0", tk.END).strip():
				messagebox.showwarning("Empty", "Content required.", parent=self.window)
				return
			self.result_id = self.store.create_payload(
				e_name.get(), e_cat.get() or "other", e_plat.get() or "any",
				t_content.get("1.0", tk.END).rstrip(), e_tags.get(),
			)
			self.window.destroy()
		btns = tk.Frame(self.window, bg=BG)
		btns.grid(row=5, column=0, columnspan=2, pady=8)
		ttk.Button(btns, text="Save", command=save).pack(side="left", padx=4)
		ttk.Button(btns, text="Cancel", command=self.window.destroy).pack(side="left", padx=4)
		self.window.bind("<Return>", lambda e: save())
		self.window.bind("<Escape>", lambda e: self.window.destroy())


def main():
	root = tk.Tk()
	try:
		root.tk.call("tk", "scaling", 1.4)
	except Exception:
		pass
	App(root)
	root.mainloop()


if __name__ == "__main__":
	main()